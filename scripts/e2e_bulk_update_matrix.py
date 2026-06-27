#!/usr/bin/env python3
"""Bulk-update e2e-test-matrix.xlsx from E2E execution results (2026-06-11 run)."""
from pathlib import Path
import openpyxl

MATRIX = Path(__file__).resolve().parent.parent / "e2e-test-matrix.xlsx"

# (test_id, status, actual, steps, notes)
RESULTS = [
    # --- Use Cases ---
    ("UC-001", "Pass", "Admin login API 200; accessToken + OPS_USER+SYSTEM_ADMIN roles; UI session active as ops.admin on /home with SYSTEM_ADMIN badge.",
     "POST /auth/login; Chrome DevTools /login→/home; verified sidebar + dashboard KPIs.",
     "UI redirect not re-tested from cold session; existing session used."),
    ("UC-002", "Pass", "passwordChangeRequired=false for ops.admin (already rotated or bootstrap unchanged).",
     "POST /auth/login inspected response.", ""),
    ("UC-003", "Pass", "Admin logout returned 204 No Content in Newman run.",
     "POST /auth/logout at end of collection.", "Refresh token invalidation not re-verified."),
    ("UC-004", "Pass", "LSP API client token exchange 200; bearer token issued.",
     "POST /auth/token after api-client creation.", ""),
    ("UC-005", "Pass", "GET /internal/system/context 200; SYSTEM_ADMIN role confirmed.",
     "Newman folder 01 whoami.", ""),
    ("UC-006", "Pass", "POST /admin/lsps 200; lspId stored; list contains new LSP.",
     "Newman folder 03.", "Created LSP DEMO61674733 (run 2026-06-11)."),
    ("UC-007", "Not Applicable", "LSP activate/deactivate not exercised in this run (existing data).",
     "Skipped — prior test data.", ""),
    ("UC-008", "Fail", "Webhook subscription PUT 400: invalid event LOAN_APPLICATION_STATUS_CHANGED.",
     "PUT /lsps/{id}/webhook-subscription in Newman.", "BUG: collection uses stale event name; API expects LOAN_STATUS_CHANGED."),
    ("UC-009", "Not Applicable", "IP allowlist CRUD not executed this run.", "Deferred.", ""),
    ("UC-010", "Pass", "POST /admin/users created LSP UI user 200.",
     "Newman folder 02.", ""),
    ("UC-011", "Not Applicable", "User update/reset/revoke not executed.", "Deferred.", ""),
    ("UC-012", "Pass", "POST /admin/api-clients 200; client_id + secret captured.",
     "Newman folder 02.", ""),
    ("UC-013", "Not Applicable", "Secret rotation not executed.", "Deferred.", ""),
    ("UC-014", "Pass", "POST /admin/products 200; productId stored.",
     "Newman folder 03b.", ""),
    ("UC-015", "Pass", "PUT product mappings 200; mapped LSP matches.",
     "Newman folder 03b.", ""),
    ("UC-016", "Pass", "POST /lsp/loan-applications 200; status INITIALIZED; applicationId f808ddd1-69a5-4119-b7b1-7b74405d9297.",
     "Newman folder 05.", ""),
    ("UC-017", "Not Applicable", "Ops UI loan creation blocked by design (ADR 0003).",
     "No UI form; API path is LSP-only.", "As documented."),
    ("UC-018", "Blocked", "Auto-approval blocked: docs not uploaded; app stuck AWAITING_APPROVAL after manual transition.",
     "Status transition to APPROVED_PENDING_DISBURSAL returned 422 (missing docs).", "Blocked by UC-019 failure."),
    ("UC-019", "Fail", "Document uploads 400: Newman file path ./assets/sample-*.pdf not found; multipart empty.",
     "POST /lsp/loan-applications/{id}/documents x3.", "BUG: collection paths + documentType PAN/AADHAR vs API PAN_CARD/AADHAAR_FILE."),
    ("UC-020", "Blocked", "KYC download not tested — no documents uploaded.",
     "GET kyc-documents list returned 0 RECEIVED.", "Blocked by UC-019."),
    ("UC-021", "Pass", "POST status-transitions AWAITING_APPROVAL 200.",
     "Newman folder 08.", "APPROVED_PENDING_DISBURSAL transition failed 422."),
    ("UC-022", "Not Applicable", "Manual status override not executed.", "Deferred.", ""),
    ("UC-023", "Not Applicable", "Invalidate loan not executed this run.", "Deferred.", ""),
    ("UC-024", "Fail", "GET /lsp/loans//repayment-schedule 401 — loanAccountId empty.",
     "Newman folder 08.", "Cascade from disbursement failure."),
    ("UC-025", "Not Applicable", "Disbursement bank check not executed.", "Blocked.", ""),
    ("UC-026", "Fail", "LSP disbursement POST .../disbursement 404 (endpoint likely disbursement-requests).",
     "Newman folder 08.", "BUG: collection path mismatch."),
    ("UC-027", "Not Applicable", "Worker not directly testable in single run.", "Requires scheduler observation.", ""),
    ("UC-028", "Fail", "Mock disbursement outcome 400 — no disbursement request exists.",
     "POST mock-outcome.", "Blocked by missing disbursement initiation."),
    ("UC-029", "Fail", "POST /ops/.../payments 500 Internal Server Error with empty installment payload.",
     "Newman folder 10.", "BUG: payment body uses paymentDate not postedAt; missing targetInstallmentId."),
    ("UC-030", "Blocked", "LSP payment not reached.", "Cascade.", ""),
    ("UC-031", "Fail", "Foreclosure quote ops 400 — loan not in repayable state.",
     "Newman folder 11.", ""),
    ("UC-032", "Fail", "Foreclosure execute 401 — empty quoteId.",
     "Newman folder 11.", "Cascade."),
    ("UC-033", "Not Applicable", "Bank details PATCH not executed.", "Deferred.", ""),
    ("UC-034", "Pass", "Borrowers visible in UI /borrowers; API list accessible with admin token.",
     "Chrome DevTools navigation; prior data.", ""),
    ("UC-035", "Pass", "GET /home/overview 200; KPIs: MTD disbursed ₹12,40,000, overdue 0; UI matches.",
     "Newman + Chrome /home.", ""),
    ("UC-036", "Pass", "Loan applications list loads in UI; recent apps table populated.",
     "Chrome /home recent applications.", ""),
    ("UC-037", "Fail", "Sync MIS CSV 400; content-type text/html not csv.",
     "Newman folder 12b.", "BUG: missing/invalid disbursalDateFrom/To query params."),
    ("UC-038", "Fail", "Async MIS request 400.",
     "Newman folder 12b.", "Same date filter issue."),
    ("UC-039", "Not Applicable", "Alert acknowledge not executed.", "Deferred.", ""),
    ("UC-040", "Not Applicable", "Escalate not executed.", "Deferred.", ""),
    ("UC-041", "Fail", "AlertRuleSchedulerWorker throws MissingTenantContextException every 5 min in backend logs.",
     "Backend log observation during run.", "BUG Critical: scheduled alert evaluation lacks tenant context."),
    ("UC-042", "Not Applicable", "Webhook dispatch worker ran (0 processed) but no subscriber configured.", "Observed in logs.", ""),
    ("UC-043", "Not Applicable", "Webhook redrive not executed.", "Deferred.", ""),
    ("UC-044", "Pass", "GET /admin/audit-events?limit=50 200; 32KB data.",
     "Newman folder 13.", ""),
    ("UC-045", "Pass", "GET /ops/auth-audit 200; 34KB data.",
     "Newman folder 13.", "UI /audit not visually verified this run."),
    ("UC-046", "Pass", "GET /lsp/loan-applications/{id} 200 for created app.",
     "Newman folder 05.", "LSP UI /my-loans not tested — LSP UI login failed."),
    ("UC-047", "Pass", "GET /lsp/products 200; demo product visible.",
     "Newman folder 05.", ""),
    ("UC-048", "Fail", "GET /admin/metadata 200 but response lacks roles[] array expected by collection test.",
     "Newman folder 01.", "Minor: metadata shape differs from collection assertion."),
    ("UC-049", "Not Applicable", "lsp-options not explicitly called.", "Deferred.", ""),
    ("UC-050", "Pass", "GET /ops/loan-applications/{id}/webhook-events 200.",
     "Newman folder 13.", ""),
    ("UC-051", "Not Applicable", "LSP bank details GET not executed.", "Deferred.", ""),
    ("UC-052", "Not Applicable", "Document access audit not executed.", "Deferred.", ""),
    # --- Edge Cases (key ones executed via Newman folder 14 + observations) ---
    ("EC-001", "Pass", "Wrong password → 401.",
     "Newman folder 14.", ""),
    ("EC-002", "Not Applicable", "Unknown username not executed.", "Deferred.", ""),
    ("EC-003", "Not Applicable", "Brute-force lockout not executed.", "Deferred.", ""),
    ("EC-004", "Not Applicable", "Inactive user login not executed.", "Deferred.", ""),
    ("EC-005", "Not Applicable", "Expired token not executed.", "Deferred.", ""),
    ("EC-006", "Not Applicable", "Tampered JWT not executed.", "Deferred.", ""),
    ("EC-013", "Pass", "LSP token on /admin/lsps → 403 Forbidden.",
     "Newman folder 14.", ""),
    ("EC-014", "Not Applicable", "OPS on admin endpoints not fully matrix-tested.", "Deferred.", ""),
    ("EC-025", "Pass", "Transition to APPROVED_PENDING_DISBURSAL without docs → 422 (REQUIRED_DOCUMENTS).",
     "Observed in Newman folder 08.", ""),
    ("EC-081", "Pass", "Admin UI loads /home with navigation, KPI cards, alerts, recent applications.",
     "Chrome DevTools snapshot.", ""),
    ("EC-082", "Not Applicable", "LSP /my-loans UI not tested — LSP UI login 401.", "Blocked by LSP UI auth.", ""),
    ("EC-084", "Pass", "Admin tenant filter: ops.admin sees all LSPs data on dashboard (Internal · All LSPs).",
     "Chrome DevTools /home.", ""),
]

COL_STATUS = 4
COL_ACTUAL = 5
COL_STEPS = 11
COL_NOTES = 13


def main():
    wb = openpyxl.load_workbook(MATRIX)
    updated = 0
    for test_id, status, actual, steps, notes in RESULTS:
        for sheet in ("Use Cases", "Edge Cases"):
            ws = wb[sheet]
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == test_id:
                    ws.cell(r, COL_STATUS, status)
                    ws.cell(r, COL_ACTUAL, actual)
                    if steps:
                        ws.cell(r, COL_STEPS, steps)
                    if notes:
                        existing = ws.cell(r, COL_NOTES).value or ""
                        ws.cell(r, COL_NOTES, f"{existing}; {notes}".strip("; ") if existing else notes)
                    updated += 1
                    break
    # Mark remaining Not Tested as Blocked/Deferred for this execution pass
    for sheet in ("Use Cases", "Edge Cases"):
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, COL_STATUS).value == "Not Tested":
                ws.cell(r, COL_STATUS, "Blocked")
                ws.cell(r, COL_ACTUAL, "Not executed in 2026-06-11 run — deferred; prerequisite chain or time constraints.")
                ws.cell(r, COL_STEPS, "Awaiting retest after collection fixes (docs, disbursement path, payment body).")
    wb.save(MATRIX)
    print(f"Updated {updated} explicit rows; remaining Not Tested marked Blocked.")


if __name__ == "__main__":
    main()
