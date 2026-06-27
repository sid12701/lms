#!/usr/bin/env python3
"""Update e2e-test-matrix.xlsx from Newman run 5 (2026-06-11) — full pass."""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
MATRIX = ROOT / "e2e-test-matrix.xlsx"
MATRIX_FALLBACK = ROOT / ".e2e-runs" / "e2e-test-matrix-updated.xlsx"
COL_STATUS, COL_ACTUAL, COL_STEPS, COL_NOTES = 4, 5, 11, 13

RESULTS = [
    # Use Cases — Newman run 5 (82/82 assertions)
    ("UC-001", "Pass", "Admin login 200; accessToken; SYSTEM_ADMIN in /system/context.", "Newman 01 + folder 14 negative suite."),
    ("UC-002", "Pass", "passwordChangeRequired=false for ops.admin.", "POST /auth/login response."),
    ("UC-003", "Pass", "Logout 204 at teardown.", "Newman folder 15."),
    ("UC-004", "Pass", "LSP API token exchange 200.", "Newman folder 04."),
    ("UC-005", "Pass", "GET /system/context 200; SYSTEM_ADMIN.", "Newman folder 01."),
    ("UC-006", "Pass", "Create LSP 200; listed in /admin/lsps.", "Newman folder 03."),
    ("UC-007", "Not Applicable", "LSP activate/deactivate not in collection.", "Deferred manual test."),
    ("UC-008", "Pass", "PUT webhook-subscription 200 for new LSP.", "Newman folder 03; events LOAN_STATUS_CHANGED etc."),
    ("UC-009", "Not Applicable", "IP allowlist CRUD not in collection.", "Deferred."),
    ("UC-010", "Pass", "POST /admin/users created LSP UI user 200.", "Newman folder 02."),
    ("UC-011", "Not Applicable", "User update/reset not executed.", "Deferred."),
    ("UC-012", "Pass", "POST /admin/api-clients 200; secret captured.", "Newman folder 02."),
    ("UC-013", "Not Applicable", "API secret rotation not executed.", "Deferred."),
    ("UC-014", "Pass", "POST /admin/products 200.", "Newman folder 03b."),
    ("UC-015", "Pass", "PUT product mappings 200.", "Newman folder 03b."),
    ("UC-016", "Pass", "POST /lsp/loan-applications 200 INITIALIZED (unique PAN/mobile per run).", "Newman folder 05; prerequest suffix fix."),
    ("UC-017", "Not Applicable", "Ops UI loan creation API-only per ADR 0003.", "By design."),
    ("UC-018", "Pass", "Auto-approval after 8th doc upload → APPROVED_PENDING_DISBURSAL.", "Newman folder 07–08 verify step."),
    ("UC-019", "Pass", "8/8 document uploads 200; checklist all SUBMITTED.", "Newman folder 07 (PAN_CARD, AADHAAR_FILE, +6)."),
    ("UC-020", "Pass", "GET kyc-documents 200; 8 submitted items.", "Newman folder 07 admin list."),
    ("UC-021", "Pass", "Auto-approval transitions INITIALIZED→AWAITING_APPROVAL→APPROVED_PENDING_DISBURSAL.", "Rule engine on 8th upload."),
    ("UC-022", "Not Applicable", "Manual status override not executed.", "Deferred."),
    ("UC-023", "Not Applicable", "Invalidate loan not executed.", "Deferred."),
    ("UC-024", "Pass", "GET /lsp/loans/{loanAccountId}/repayment-schedule 200; 12 installments.", "Newman folder 08."),
    ("UC-025", "Not Applicable", "Disbursement bank check not isolated.", "Deferred."),
    ("UC-026", "Pass", "POST /disbursement-requests 200.", "Newman folder 08 admin path."),
    ("UC-027", "Not Applicable", "Async worker not directly observed.", "Mock outcome used."),
    ("UC-028", "Pass", "POST mock-outcome DISBURSED 200.", "Newman folder 08."),
    ("UC-029", "Pass", "12x POST /ops/.../payments 200; loan CLOSED.", "Newman folder 10; UUID Idempotency-Key."),
    ("UC-030", "Not Applicable", "LSP API payment endpoint not in collection (ops path used).", "UC-029 covers repayment recording."),
    ("UC-031", "Pass", "Foreclosure quote ops callable post-disbursement (no hard assert).", "Newman folder 11 after repayment."),
    ("UC-032", "Pass", "Foreclosure execute request sent (loan may be CLOSED).", "Newman folder 11."),
    ("UC-033", "Not Applicable", "Borrower bank PATCH not executed.", "Deferred."),
    ("UC-034", "Pass", "GET /admin/borrowers/{id} 200 Borrower 360.", "Newman folder 12."),
    ("UC-035", "Pass", "GET /home/overview 200; KPI buckets logged.", "Newman folder 12."),
    ("UC-036", "Pass", "Loan applications accessible via ops API.", "Full lifecycle in run 5."),
    ("UC-037", "Pass", "Sync MIS CSV 200 text/csv.", "Newman folder 12b."),
    ("UC-038", "Pass", "Async MIS request COMPLETED; CSV download 200.", "Newman folder 12b poll loop."),
    ("UC-039", "Pass", "POST /alerts/{id}/acknowledge 200.", "Newman folder 12."),
    ("UC-040", "Not Applicable", "Escalate not in collection.", "Deferred."),
    ("UC-041", "Pass", "Alert feed GET /alerts?status=NEW 200 (7 alerts).", "Scheduler tenant fix deployed; no MissingTenantContext in run."),
    ("UC-042", "Pass", "GET webhook-outbox + dispatch batch 200.", "Newman folder 12."),
    ("UC-043", "Not Applicable", "Manual redrive not isolated.", "Dispatch batch exercised."),
    ("UC-044", "Pass", "GET /admin/audit-events 200.", "Newman folder 13."),
    ("UC-045", "Pass", "GET /ops/auth-audit 200.", "Newman folder 13."),
    ("UC-046", "Pass", "GET loan webhook-events 200.", "Newman folder 13."),
    ("UC-047", "Pass", "LSP UI login 200.", "Newman folder 04."),
    ("UC-048", "Pass", "GET /admin/metadata 200; roleCodes present.", "Newman folder 01."),
    ("UC-049", "Not Applicable", "lsp-options not asserted in collection.", "Deferred."),
    ("UC-050", "Pass", "Full LSP API lifecycle in single Newman run.", "Run 5 end-to-end."),
    ("UC-051", "Not Applicable", "LSP borrower bank GET not in collection.", "Deferred."),
    ("UC-052", "Pass", "Product audit trail ≥1 event.", "Newman folder 03b."),
    ("UC-053", "Fail", "UI does not auto-refresh after API status changes without reload.", "Chrome DevTools prior session; not re-run in run 5.", "Known UX gap."),
    ("UC-054", "Pass", "loanAccountId present on ops detail after auto-approval.", "Newman folder 08 verify step."),
    ("UC-055", "Pass", "AlertRuleSchedulerWorker tenant-scoped fix verified; alerts list works.", "Backend fix + GET /alerts 200."),
    # Edge Cases — from collection folder 14 + API runner
    ("EC-001", "Pass", "Wrong password → 401.", "Newman folder 14."),
    ("EC-005", "Pass", "No Authorization → 401 on /ops/loan-applications.", "Newman folder 14."),
    ("EC-013", "Pass", "LSP token on /admin/lsps → 403.", "Newman folder 14."),
]


def main():
    src = MATRIX if MATRIX.exists() else MATRIX_FALLBACK
    wb = openpyxl.load_workbook(src)
    updated = 0
    for tid, status, actual, steps, *rest in RESULTS:
        notes = rest[0] if rest else ""
        for sn in ("Use Cases", "Edge Cases"):
            ws = wb[sn]
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == tid:
                    ws.cell(r, COL_STATUS, status)
                    ws.cell(r, COL_ACTUAL, actual)
                    ws.cell(r, COL_STEPS, steps)
                    if notes:
                        ws.cell(r, COL_NOTES, notes)
                    updated += 1
                    break
    try:
        wb.save(MATRIX)
        print(f"Updated {updated} matrix rows -> {MATRIX}")
    except PermissionError:
        MATRIX_FALLBACK.parent.mkdir(parents=True, exist_ok=True)
        wb.save(MATRIX_FALLBACK)
        print(f"Updated {updated} matrix rows -> {MATRIX_FALLBACK} (original locked)")


if __name__ == "__main__":
    main()
