#!/usr/bin/env python3
"""Set specific Blocked reasons on remaining edge cases in e2e-test-matrix.xlsx."""
from pathlib import Path

import openpyxl

MATRIX = Path(__file__).resolve().parent.parent / "e2e-test-matrix.xlsx"
COL_STATUS, COL_ACTUAL, COL_STEPS, COL_NOTES = 4, 5, 11, 13

# test_id -> (actual, steps, notes)
REASONS: dict[str, tuple[str, str, str]] = {
    # Auth / session / credential state
    "EC-007": (
        "Blocked: needs logout + reuse of revoked access token (session invalidation path).",
        "Not run — requires capturing refresh/session version after forced revoke.",
        "Auth session lifecycle test; Newman only exercises happy-path login/logout.",
    ),
    "EC-008": (
        "Blocked: needs API client secret rotation then token exchange within grace window.",
        "Not run — requires POST rotate-secret then immediate /auth/token.",
        "Fixture: active client with old+new secret both valid briefly.",
    ),
    "EC-009": (
        "Blocked: needs LSP set INACTIVE then API client token attempt.",
        "Not run — requires admin PATCH LSP status then /auth/token.",
        "Fixture: dedicated inactive LSP + client (not created in Newman run).",
    ),
    "EC-010": (
        "Blocked: needs IP allowlist enabled on LSP with request from non-allowed IP.",
        "Not run — local Newman calls 127.0.0.1; allowlist rules not configured.",
        "Infra: configure allowlist excluding runner IP or use proxy.",
    ),
    "EC-011": (
        "Blocked: same as EC-010 but LSP UI login path.",
        "Not run — IP allowlist + /auth/login from blocked IP.",
        "Infra prerequisite.",
    ),
    # Multi-tenant / RBAC
    "EC-012": (
        "Blocked: needs second tenant/LSP and cross-tenant resource UUID.",
        "Not run — single-tenant local seed only.",
        "Requires two isolated LSP tenants in DB.",
    ),
    "EC-015": (
        "Blocked: needs PRODUCT_ADMIN user (not created in Newman).",
        "Not run — only SYSTEM_ADMIN + LSP users exercised.",
        "Create PRODUCT_ADMIN via /admin/users then probe forbidden route.",
    ),
    "EC-016": (
        "Blocked: needs attempt to disable sole SYSTEM_ADMIN self.",
        "Not run — ops.admin is only admin exercised; risky on shared DB.",
        "Manual/sandbox: PATCH user active=false on self.",
    ),
    "EC-017": (
        "Blocked: needs admin reset password then verify mustChangePassword on next login.",
        "Not run — password reset flow not in collection.",
        "API: POST reset-password + login assertion.",
    ),
    # Auto-rejection (negative fixtures)
    "EC-018": (
        "Blocked: needs loan against INACTIVE product.",
        "Not run — Newman creates ACTIVE product only.",
        "Create product INACTIVE or deactivate after mapping.",
    ),
    "EC-019": (
        "Blocked: needs INACTIVE LSP at intake time.",
        "Not run — Newman creates ACTIVE LSP.",
        "Deactivate LSP before POST /lsp/loan-applications.",
    ),
    "EC-020": (
        "Blocked: needs inactive LSP–product mapping.",
        "Not run — mapping set ACTIVE in folder 03b.",
        "Deactivate mapping then submit application.",
    ),
    "EC-024": (
        "Blocked: needs intake payload missing required borrower fields.",
        "Not run — EC-028 covers generic validation; this targets auto-rejection reason code.",
        "POST with empty employment/income fields; assert rejection_reason_json.",
    ),
    # Idempotency / payload size
    "EC-027": (
        "Blocked: needs duplicate POST /lsp/loan-applications with same Idempotency-Key.",
        "Not run — idempotency replay not scripted.",
        "Automatable: replay same key + body; expect same applicationId.",
    ),
    "EC-032": (
        "Blocked: needs request body > server max (multipart/json).",
        "Not run — no oversized payload generator in runner.",
        "Automatable: POST >10MB body.",
    ),
    "EC-038": (
        "Blocked: needs invalidate endpoint with repeated idempotency key.",
        "Not run — invalidation flow not in Newman happy path.",
        "Automatable after loan in pre-disbursal state.",
    ),
    # Illegal status transitions (terminal states)
    "EC-033": (
        "Blocked: needs REJECTED loan then illegal transition attempt.",
        "Not run — run5 loan ends CLOSED not REJECTED for this probe.",
        "Automatable: create reject via rule failure then POST transition.",
    ),
    "EC-034": (
        "Blocked: needs CLOSED loan then illegal transition.",
        "Not run — run5 loan CLOSED but transition negative not asserted.",
        "Automatable using run5 applicationId.",
    ),
    "EC-035": (
        "Blocked: needs FORECLOSED loan then illegal transition.",
        "Not run — run5 closed via repayment not foreclosure.",
        "Automatable if foreclosure path executed first.",
    ),
    "EC-036": (
        "Blocked: needs INVALID loan state.",
        "Not run — invalidate flow not executed.",
        "Automatable via ops invalidate pre-disbursal.",
    ),
    "EC-037": (
        "Blocked: needs invalidate after DISBURSED.",
        "Not run — post-disbursement invalidate not attempted.",
        "Automatable on run5 disbursed app.",
    ),
    "EC-039": (
        "Blocked: needs manual transition without reasonCode where required.",
        "Not run — auto-approval path used instead of manual ops transition.",
        "Automatable: POST transition missing reason on approval.",
    ),
    # Document upload negatives
    "EC-041": (
        "Blocked: needs file > DocumentUploadPolicy max bytes.",
        "Not run — only valid PDFs uploaded.",
        "Automatable: generate >10MB PDF.",
    ),
    "EC-042": (
        "Blocked: needs documentType not on product checklist.",
        "Not run — only standard 8 types uploaded.",
        "Automatable if API accepts arbitrary enum value.",
    ),
    "EC-044": (
        "Blocked: needs second LSP tenant downloading another tenant's document.",
        "Not run — single LSP in run.",
        "Multi-tenant fixture required.",
    ),
    "EC-045": (
        "Blocked: needs download of document type never uploaded.",
        "Not run — GET download for missing type not called.",
        "Automatable on run5 app.",
    ),
    # Disbursement negatives
    "EC-046": (
        "Blocked: needs disbursement on INITIALIZED/AWAITING (non-eligible).",
        "Not run — happy path reaches APPROVED_PENDING_DISBURSAL first.",
        "Automatable: POST disbursement-requests on INITIALIZED app.",
    ),
    "EC-047": (
        "Blocked: needs disbursement without schedule (if applicable).",
        "Not run — schedule auto-generated on disburse in mock mode.",
        "May be N/A depending on product config.",
    ),
    "EC-048": (
        "Blocked: needs bank account mismatch vs borrower on file.",
        "Not run — consistent bank details in Newman payload.",
        "Automatable: alter bank fields before disbursement.",
    ),
    "EC-049": (
        "Blocked: needs disbursement adapter retry exhaustion.",
        "Not run — mock adapter succeeds first time.",
        "Requires mock failure configuration or worker observation.",
    ),
    "EC-050": (
        "Blocked: needs mock-outcome with test mode disabled.",
        "Not run — local profile uses mock disbursement.",
        "Env: run with non-mock/disabled test flag.",
    ),
    # Payment negatives (EC-051 executed → Fail 500)
    "EC-052": (
        "Blocked: needs payment body omitting targetInstallmentId.",
        "Not run — Newman supplies full payment body.",
        "Automatable API negative.",
    ),
    "EC-053": (
        "Blocked: needs partial installment amount (< due).",
        "Not run — full EMI amounts paid in run5.",
        "Automatable on UNDER_REPAYMENT loan.",
    ),
    "EC-054": (
        "Blocked: needs second payment to same PAID installment.",
        "Not run — 12 distinct installments paid once each.",
        "Automatable: replay payment to installment 1.",
    ),
    "EC-055": (
        "Blocked: needs payment idempotency key replay (expect same receipt).",
        "Not run — new UUID per installment in Newman.",
        "Automatable: repeat POST with same Idempotency-Key.",
    ),
    "EC-056": (
        "Blocked: needs invalid payment channel enum.",
        "Not run — channel=NEFT only.",
        "Automatable API negative.",
    ),
    "EC-057": (
        "Blocked: needs payment on CLOSED/FORECLOSED loan.",
        "Not run — run5 ends CLOSED; payment-after-close not probed.",
        "Automatable using run5 closed applicationId.",
    ),
    # Foreclosure negatives
    "EC-058": (
        "Blocked: needs foreclosure quote on INITIALIZED/APPROVED loan.",
        "Not run — quote attempted post-disburse in folder 11 only.",
        "Automatable on pre-disburse app.",
    ),
    "EC-059": (
        "Blocked: needs execute with expired quote.",
        "Not run — no quote TTL wait/expire scripted.",
        "Automatable: create quote, wait/advance clock, execute.",
    ),
    "EC-060": (
        "Blocked: needs foreclosure execute on CLOSED loan.",
        "Not run — run5 CLOSED via repayment; execute not asserted.",
        "Automatable on closed applicationId.",
    ),
    # Webhook integration / security
    "EC-061": (
        "Blocked: needs webhook URL with DNS failure.",
        "Not run — webhook.site optional; no invalid DNS URL configured.",
        "Infra: subscribe https://nonexistent.invalid domain.",
    ),
    "EC-062": (
        "Blocked: needs mock HTTP server returning 5xx.",
        "Not run — no local webhook receiver in test harness.",
        "Infra: webhook.site or mock server returning 500.",
    ),
    "EC-063": (
        "Blocked: needs subscriber returning 4xx.",
        "Not run — same as EC-062.",
        "Infra prerequisite.",
    ),
    "EC-064": (
        "Blocked: needs OPS_USER calling admin redrive endpoint.",
        "Not run — Newman uses SYSTEM_ADMIN for dispatch.",
        "Create OPS_USER and POST redrive.",
    ),
    "EC-065": (
        "Blocked: needs inbound webhook payload verification without signature.",
        "Not run — outbound-only exercised.",
        "Requires subscriber-side signature validation test.",
    ),
    "EC-066": (
        "Blocked: needs SSRF webhook URL (internal IP/metadata).",
        "Not run — security negative not attempted on local.",
        "PUT subscription with http://169.254.169.254 etc.; expect rejection.",
    ),
    "EC-105": (
        "Blocked: needs ordered webhook events per aggregate asserted.",
        "Not run — outbox listed but strict ordering not verified.",
        "Automatable: compare sequence nos on loan webhook-events.",
    ),
    # MIS / reports
    "EC-067": (
        "Blocked: needs MIS preview row inspection for unmasked aadhaar.",
        "Not run — preview returned 200 but PII mask not manually inspected.",
        "Automatable: parse preview JSON for 12-digit aadhaar pattern.",
    ),
    "EC-068": (
        "Blocked: needs large portfolio dataset for CSV performance.",
        "Not run — single-loan MIS in run5.",
        "Seed many loans or use staging data volume.",
    ),
    "EC-069": (
        "Blocked: needs LSP/OPS user without report permission.",
        "Not run — admin token used for MIS.",
        "Create restricted user; GET report expect 403.",
    ),
    "EC-070": (
        "Blocked: needs MIS date range with zero rows.",
        "Not run — run5 uses 2026-01-01..2026-12-31 with data.",
        "Automatable: future date range query.",
    ),
    # Borrower / bank
    "EC-071": (
        "Blocked: needs LSP A querying borrower belonging to LSP B.",
        "Not run — single LSP tenant.",
        "Multi-tenant fixture.",
    ),
    "EC-072": (
        "Blocked: needs PATCH bank with invalid IFSC.",
        "Not run — bank update endpoint not in Newman.",
        "Automatable API negative.",
    ),
    # Rate limits
    "EC-074": (
        "Blocked: needs >N login attempts same IP+username.",
        "Not run — rate limit threshold not hammered.",
        "Load script or loop >limit POST /auth/login.",
    ),
    "EC-075": (
        "Blocked: needs burst /auth/token per client_id.",
        "Not run — single token exchange in run.",
        "Load script prerequisite.",
    ),
    "EC-076": (
        "Blocked: needs burst POST /lsp/loan-applications.",
        "Not run — one application per Newman run.",
        "Load script prerequisite.",
    ),
    # Alert edge cases
    "EC-077": (
        "Blocked: needs POST acknowledge on random alert UUID.",
        "Not run — Newman acknowledges real alert only.",
        "Automatable: POST /alerts/{random}/acknowledge.",
    ),
    "EC-078": (
        "Blocked: needs acknowledge note >500 chars.",
        "Not run — default note length in collection.",
        "Automatable API negative.",
    ),
    "EC-079": (
        "Blocked: needs double-acknowledge same alert.",
        "Not run — single acknowledge in run5.",
        "Automatable: acknowledge twice.",
    ),
    "EC-080": (
        "Blocked: needs duplicate alert condition on same loan.",
        "Not run — scheduler/rule dedupe not provoked.",
        "Requires triggering same rule twice without ack.",
    ),
    # UI (Chrome DevTools)
    "EC-083": (
        "Blocked: browser refresh session persistence — UI test.",
        "Not run — no Chrome DevTools pass in this session.",
        "Chrome: login, refresh /loan-applications/{id}, assert still authenticated.",
    ),
    "EC-085": (
        "Blocked: console error scan on page load — UI test.",
        "Not run — no DevTools console audit.",
        "Chrome: navigate key routes, assert console clean.",
    ),
    "EC-086": (
        "Blocked: ErrorState UI on failed fetch — UI test.",
        "Not run — requires simulating offline/500 in browser.",
        "Chrome: block network or mock 500.",
    ),
    "EC-087": (
        "Blocked: empty list UI states — UI test.",
        "Not run — DB has seed data; empty filters not exercised in UI.",
        "Chrome: filter to zero results, assert empty state component.",
    ),
    "EC-088": (
        "Blocked: pagination last-page boundaries — UI test.",
        "Not run — audit/list pagination not clicked through UI.",
        "Chrome: /audit or /loan-applications pagination.",
    ),
    "EC-089": (
        "Blocked: filter combinations on /loan-applications — UI test.",
        "Not run — UI filters not systematically combined.",
        "Chrome DevTools interaction test.",
    ),
    "EC-090": (
        "Blocked: sensitive-data reveal toggle — UI test.",
        "Not run — PII mask toggle not exercised in browser.",
        "Chrome: toggle mask on borrower/loan detail.",
    ),
    "EC-096": (
        "Blocked: UI status label vs API/DB — UI test.",
        "Not run — no side-by-side UI+API compare this session.",
        "Chrome detail page vs GET ops application.",
    ),
    "EC-097": (
        "Blocked: webhook event count in UI vs outbox API.",
        "Not run — UI webhook tab not compared to API.",
        "Chrome loan detail webhooks tab.",
    ),
    "EC-111": (
        "Blocked: UI does not poll after external API status change.",
        "Not run — same as UC-053; requires API change without reload.",
        "Chrome: change status via API, observe stale UI.",
    ),
    # Audit
    "EC-093": (
        "Blocked: needs audit query with unknown stream enum.",
        "Not run — default streams only.",
        "Automatable: GET audit-events?stream=INVALID.",
    ),
    "EC-094": (
        "Blocked: needs INTAKE audit payload aadhaar mask inspection.",
        "Not run — audit JSON not scanned for raw aadhaar.",
        "Automatable: filter INTAKE events, regex check.",
    ),
    "EC-095": (
        "Blocked: needs document download then document-access-audit row.",
        "Not run — uploads done; download+audit link not verified.",
        "Automatable: GET document download then GET document-access-audits.",
    ),
    "EC-110": (
        "Blocked: needs correlationId trace across audit explorer.",
        "Not run — correlation from Newman logged but not traced in UI/API.",
        "Automatable: grep audit by X-Correlation-Id from run5.",
    ),
    # Lifecycle assertions (partially covered by run5)
    "EC-098": (
        "Blocked: explicit assert first payment → UNDER_REPAYMENT.",
        "Not run — run5 paid installments but transition row not inspected.",
        "Automatable: GET status-transitions after payment 1.",
    ),
    "EC-099": (
        "Blocked: explicit assert last payment → CLOSED.",
        "Not run — run5 asserts CLOSED on detail only.",
        "Automatable: verify final transition record.",
    ),
    "EC-100": (
        "Blocked: rejection_reason_json on REJECTED transition.",
        "Not run — no REJECTED loan in run5 happy path.",
        "Automatable via auto-reject fixture (EC-018–020).",
    ),
    # Regression / ADR / business rules
    "EC-102": (
        "Blocked: local 401 loop regression (#89) — profile-specific filter.",
        "Not run — TestTenantContextRestoreFilter @Profile test only.",
        "Requires test profile or documented repro from issue #89.",
    ),
    "EC-103": (
        "Blocked: processing fee deduction at disbursement (ADR 0004).",
        "Not run — disbursement amount vs net proceeds not reconciled.",
        "Automatable: compare disbursedAmount - processingFee in API response.",
    ),
    "EC-106": (
        "Blocked: bank holder-name fuzzy match on disbursement.",
        "Not run — exact name match in Newman payload.",
        "Automatable: slight name mismatch on accountHolderName.",
    ),
    "EC-107": (
        "Blocked: schedule submission validation errors.",
        "Not run — schedule auto-generated; manual submit not tested.",
        "Automatable if manual schedule API exists.",
    ),
    "EC-108": (
        "Blocked: schedule update after disbursement blocked.",
        "Not run — post-disburse schedule mutation not attempted.",
        "Automatable on DISBURSED loan.",
    ),
    "EC-109": (
        "Blocked: concurrent disbursement attempts.",
        "Not run — single sequential disbursement in Newman.",
        "Automatable: parallel POST disbursement-requests.",
    ),
}


def main():
    wb = openpyxl.load_workbook(MATRIX)
    ws = wb["Edge Cases"]
    updated = 0
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(r, 1).value
        if ws.cell(r, 4).value != "Blocked" or tid not in REASONS:
            continue
        actual, steps, notes = REASONS[tid]
        ws.cell(r, COL_ACTUAL, actual)
        ws.cell(r, COL_STEPS, steps)
        ws.cell(r, COL_NOTES, notes)
        updated += 1
    wb.save(MATRIX)
    print(f"Updated {updated} blocked edge case rows with specific reasons.")


if __name__ == "__main__":
    main()
