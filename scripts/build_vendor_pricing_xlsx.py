# -*- coding: utf-8 -*-
"""
Build a formula-driven vendor pricing comparison workbook: Yubi (Vendor A) vs M2P (Vendor B)
for Bhawana Finance's LMS, on a EUR 100M/yr (annual disbursal) portfolio.

Layout follows the pricing block of docs/Vendor-Scorecard(Vendor Scorecard - BLANK).csv:
INR table, EUR table, first-year investment, plus a 3-year TCO and sensitivity.

All blue/yellow input cells are editable; every other number is a live formula.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

OUT = r"D:\Desktop-New\Folders\LMS\docs\Vendor-Pricing-Comparison-Yubi-vs-M2P.xlsx"

# ---------- palette ----------
NAVY      = "1F3864"
BLUE      = "2E5496"
LIGHTBLUE = "D6E0F0"
YELLOW    = "FFF2CC"
INPUTFILL = "FFF7D9"
GREY      = "F2F2F2"
GREEN     = "E2EFDA"
RED       = "FCE4D6"
WHITE     = "FFFFFF"

# ---------- number formats ----------
INR = '[>=10000000]"₹ "#,##,##,##0;[>=100000]"₹ "#,##,##0;"₹ "#,##0'
EUR = '"€ "#,##0'
PCT4 = '0.0000%'
PCT2 = '0.00%'
NUM = '#,##0'
NUM2 = '#,##0.00'

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(c, *, bold=False, size=10, color="000000", fill=None, align=None,
               wrap=False, fmt=None, border=True, italic=False):
    c.font = Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    al = {}
    if align: al["horizontal"] = align
    al["vertical"] = "center"
    if wrap: al["wrap_text"] = True
    c.alignment = Alignment(**al)
    if fmt: c.number_format = fmt
    if border: c.border = BORDER
    return c

def put(ws, ref, value, **kw):
    c = ws[ref]
    c.value = value
    style_cell(c, **kw)
    return c

def title_row(ws, ref_range, text):
    ws.merge_cells(ref_range)
    first = ref_range.split(":")[0]
    put(ws, first, text, bold=True, size=13, color=WHITE, fill=NAVY, align="left", border=False)
    for row in ws[ref_range]:
        for c in row:
            c.fill = PatternFill("solid", fgColor=NAVY)

def section(ws, ref_range, text):
    ws.merge_cells(ref_range)
    first = ref_range.split(":")[0]
    put(ws, first, text, bold=True, size=10, color=WHITE, fill=BLUE, align="left")
    for row in ws[ref_range]:
        for c in row:
            c.fill = PatternFill("solid", fgColor=BLUE)
            c.border = BORDER

wb = openpyxl.Workbook()
wb.remove(wb.active)

# =====================================================================================
# SHEET 1: ASSUMPTIONS
# =====================================================================================
A = wb.create_sheet("Assumptions")
A.sheet_properties.tabColor = YELLOW
A.column_dimensions["A"].width = 42
A.column_dimensions["B"].width = 20
A.column_dimensions["C"].width = 62
for col in "DEFG":
    A.column_dimensions[col].width = 12

title_row(A, "A1:C1", "ASSUMPTIONS & INPUTS  —  edit the shaded cells; everything else recalculates")
put(A, "A2", "Yubi vs M2P  ·  Bhawana Finance LMS  ·  €100M/yr portfolio (annual disbursal)",
    italic=True, size=9, border=False)

section(A, "A4:C4", "GLOBAL INPUTS  (editable)")
rows = [
    ("FX rate  —  INR per €1", 90, "Your booking/hedge rate. Template used 100. Drives every € figure."),
    ("Annual disbursal  (€)", 100000000, "€100M/yr = total disbursed per year (the agreed portfolio basis)."),
    ("Annual disbursal  (₹)", "=B6*B5", "Auto = annual disbursal € × FX."),
    ("Average loan tenor  (months)", 6, "Agreed 3–6 mo. Bridges disbursal→AUM for M2P. Try 3 and 6."),
    ("Avg outstanding factor (book ÷ principal)", 1.00, "1.0 = bullet (principal on book full tenor). ~0.55 = amortising EMI book."),
    ("Number of LSPs / partners live", 3, "Yubi: first 3 free. M2P floor jumps to ₹4L once >1 LSP."),
    ("Yubi free LSPs (no partnership fee)", 3, "Yubi waives partnership fee for the first 3 LSPs."),
    ("Time horizon  (years)", 3, "Drives 3-year TCO. M2P has a 36-month lock-in."),
    ("Include M2P one-time in totals?  (1=yes / 0=no)", 1, "M2P ₹55L is stated 'Paid by Meesho' (LSP). 0 = Bhawana-net view."),
    ("Yubi change-request man-days (Year 1)", 0, "₹20,000/man-day. Optional; default 0 (billed on actuals)."),
    ("M2P customization man-days (Year 1)", 0, "₹17,000/man-day. Optional; default 0 (billed on actuals)."),
]
r = 5
for label, val, note in rows:
    put(A, f"A{r}", label, bold=True, align="left")
    editable = not (isinstance(val, str) and val.startswith("="))
    put(A, f"B{r}", val, align="center", fill=INPUTFILL if editable else GREY,
        fmt=(INR if r == 7 else (EUR if r == 6 else (NUM2 if r == 9 else NUM))))
    put(A, f"C{r}", note, italic=True, size=9, align="left", wrap=True)
    r += 1

section(A, "A17:C17", "DERIVED VALUES  (auto-calculated)")
drows = [
    ("Monthly disbursal  (€)", "=B6/12", EUR),
    ("Monthly disbursal  (₹)", "=B7/12", INR),
    ("Average AUM / book  (€)", "=B6*(B8/12)*B9", EUR),
    ("Average AUM / book  (₹)", "=B7*(B8/12)*B9", INR),
    ("Average AUM  (₹ Cr)", "=B21/10000000", NUM2),
    ("M2P monthly minimum floor  (₹)", '=IF(B10>1,400000,150000)', INR),
]
r = 18
for label, val, fmt in drows:
    put(A, f"A{r}", label, bold=True, align="left", fill=GREY)
    put(A, f"B{r}", val, align="center", fmt=fmt, fill=GREY)
    note = {
        18: "= annual ÷ 12.",
        19: "= annual ÷ 12.",
        20: "= disbursal × (tenor/12) × outstanding factor.",
        21: "= disbursal × (tenor/12) × outstanding factor.",
        22: "Used for M2P fill-a-tier SaaS (1 Cr = ₹10,000,000).",
        23: "₹4L multi-LSP / ₹1.5L Meesho-only, whichever > computed SaaS.",
    }[r]
    put(A, f"C{r}", note, italic=True, size=9, align="left", wrap=True)
    r += 1

A.freeze_panes = "A3"

# cell reference helpers
ASM = "Assumptions"
fx        = f"{ASM}!$B$5"
disbEUR   = f"{ASM}!$B$6"
disbINR   = f"{ASM}!$B$7"
tenor     = f"{ASM}!$B$8"
factor    = f"{ASM}!$B$9"
LSPs      = f"{ASM}!$B$10"
freeLSP   = f"{ASM}!$B$11"
horizon   = f"{ASM}!$B$12"
inclM2P   = f"{ASM}!$B$13"
yubiCRmd  = f"{ASM}!$B$14"
m2pCustMd = f"{ASM}!$B$15"
mDisbEUR  = f"{ASM}!$B$18"
mDisbINR  = f"{ASM}!$B$19"
aumEUR    = f"{ASM}!$B$20"
aumINR    = f"{ASM}!$B$21"
aumCr     = f"{ASM}!$B$22"
m2pFloor  = f"{ASM}!$B$23"

# =====================================================================================
# SHEET 2: VENDOR A - YUBI
# =====================================================================================
Y = wb.create_sheet("Vendor A - Yubi")
Y.sheet_properties.tabColor = BLUE
widths = {"A": 40, "B": 26, "C": 18, "D": 16, "E": 46}
for c, w in widths.items():
    Y.column_dimensions[c].width = w

title_row(Y, "A1:E1", "VENDOR A  —  YUBI  (CredAvenue Pvt Ltd)   ·   LMS Commercial   ·   dated 15-Jun-2026")
hdr = ["Fee item", "Frequency / basis", "Amount (₹)", "Amount (€)", "Notes"]
for i, h in enumerate(hdr):
    put(Y, f"{get_column_letter(i+1)}3", h, bold=True, color=WHITE, fill=NAVY, align="center")

def yrow(r, item, basis, inr_formula, note, fmt=INR, bold=False, fill=None, eur=True):
    put(Y, f"A{r}", item, bold=bold, align="left", fill=fill)
    put(Y, f"B{r}", basis, align="center", size=9, fill=fill)
    put(Y, f"C{r}", inr_formula, align="right", fmt=fmt, bold=bold, fill=fill)
    if eur:
        put(Y, f"D{r}", f"=C{r}/{fx}", align="right", fmt=EUR, bold=bold, fill=fill)
    else:
        put(Y, f"D{r}", "", fill=fill)
    put(Y, f"E{r}", note, italic=True, size=9, align="left", wrap=True, fill=fill)

section(Y, "A4:E4", "ONE-TIME FEES  (paid by Bhawana Finance)")
yrow(5, "Onboarding / system integration", "one-time", 5000000, "₹75L list price − ₹25L flat discount. Invoiced on finalisation.")
yrow(6, "Yu – Reporting fee", "one-time", 1500000, "Valid for the entire association. Invoiced with onboarding.")
yrow(7, "Alert Management fee", "one-time", 1000000, "Valid for the entire association. Invoiced with onboarding.")
yrow(8, "Partnership fee (LSPs beyond free tier)", "one-time per LSP > 3", f"=500000*MAX(0,{LSPs}-{freeLSP})",
     "First 3 LSPs waived; ₹5L per LSP thereafter, at each go-live.")
yrow(9, "Subtotal — one-time", "", "=SUM(C5:C8)", "", bold=True, fill=LIGHTBLUE)

section(Y, "A11:E11", "RECURRING FEES  (per year)")
yrow(12, "Annual Maintenance Fee (AMC)", "per year", 1000000, "Year-1 invoiced upfront with onboarding; then annually.")
yrow(13, "Platform fee — 0.20% of disbursal", "per year", f"=0.002*{disbINR}",
     "0.20%/mo on monthly disbursals = 0.20% × annual disbursal. Tenor-independent.")
yrow(14, "Subtotal — recurring / year", "", "=SUM(C12:C13)", "", bold=True, fill=LIGHTBLUE)

section(Y, "A16:E16", "VARIABLE / BILLED ON ACTUALS  (excluded from totals unless man-days entered)")
put(Y, "A17", "AA / other 3rd-party API billing", align="left")
put(Y, "B17", "on actuals", align="center", size=9)
put(Y, "C17", "Pass-through", align="center", size=9, italic=True)
put(Y, "D17", "Pass-through", align="center", size=9, italic=True)
put(Y, "E17", "Excluded — depends on consumption.", italic=True, size=9, align="left", wrap=True)
yrow(18, "Change-request development", "₹20,000 / man-day", f"=20000*{yubiCRmd}",
     "Year-1 estimate from man-days input. 8-hr day per resource.")

section(Y, "A20:E20", "TOTALS")
yrow(21, "YEAR 1  (one-time + recurring + CR)", "Year 1", "=C9+C14+C18", "", bold=True, fill=GREEN)
yrow(22, "Year 2  (recurring)", "Year 2", "=C14", "", fill=GREY)
yrow(23, "Year 3  (recurring)", "Year 3", "=C14", "", fill=GREY)
yrow(24, "3-YEAR TCO", "= horizon", f"=C9+C14*{horizon}+C18", "One-time + recurring × years (+ CR Yr1).", bold=True, fill=GREEN)
Y.freeze_panes = "A4"

YUBI = "'Vendor A - Yubi'"
y_one_inr, y_one_eur = f"{YUBI}!$C$9", f"{YUBI}!$D$9"
y_rec_inr, y_rec_eur = f"{YUBI}!$C$14", f"{YUBI}!$D$14"
y_amc_inr, y_amc_eur = f"{YUBI}!$C$12", f"{YUBI}!$D$12"
y_var_inr, y_var_eur = f"{YUBI}!$C$13", f"{YUBI}!$D$13"
y_y1_inr,  y_y1_eur  = f"{YUBI}!$C$21", f"{YUBI}!$D$21"
y_tco_inr, y_tco_eur = f"{YUBI}!$C$24", f"{YUBI}!$D$24"

# =====================================================================================
# SHEET 3: VENDOR B - M2P
# =====================================================================================
M = wb.create_sheet("Vendor B - M2P")
M.sheet_properties.tabColor = "C00000"
for c, w in widths.items():
    M.column_dimensions[c].width = w

title_row(M, "A1:E1", "VENDOR B  —  M2P FINTECH (Finflux)   ·   LoS+BRE+LMS Commercial   ·   dated 24-Jun-2026")
for i, h in enumerate(hdr):
    put(M, f"{get_column_letter(i+1)}3", h, bold=True, color=WHITE, fill=NAVY, align="center")

def mrow(r, item, basis, inr_formula, note, fmt=INR, bold=False, fill=None, eur=True):
    put(M, f"A{r}", item, bold=bold, align="left", fill=fill)
    put(M, f"B{r}", basis, align="center", size=9, fill=fill)
    put(M, f"C{r}", inr_formula, align="right", fmt=fmt, bold=bold, fill=fill)
    if eur:
        put(M, f"D{r}", f"=C{r}/{fx}", align="right", fmt=EUR, bold=bold, fill=fill)
    else:
        put(M, f"D{r}", "", fill=fill)
    put(M, f"E{r}", note, italic=True, size=9, align="left", wrap=True, fill=fill)

section(M, "A4:E4", "ONE-TIME FEES")
mrow(5, "Implementation (LoS + BRE + LMS setup)", "one-time @ LOI/PO", 5500000,
     "Quote states 'Paid by Meesho' (the LSP), not Bhawana. 100% at LOI/PO.")
mrow(6, "One-time counted in totals (toggle)", "toggle-applied", f"=IF({inclM2P}=1,C5,0)",
     "Set Assumptions 'Include M2P one-time' = 0 for the Bhawana-net view.", bold=True, fill=LIGHTBLUE)

section(M, "A8:E8", "MONTHLY SaaS FEE  —  fill-a-tier (marginal slabs) on average AUM")
# tier sub-header
put(M, "A9", "AUM slab (₹ Cr)", bold=True, color=WHITE, fill=BLUE, align="center")
put(M, "B9", "Marginal rate / month", bold=True, color=WHITE, fill=BLUE, align="center")
put(M, "C9", "Monthly fee (₹)", bold=True, color=WHITE, fill=BLUE, align="center")
put(M, "D9", "AUM in slab (₹ Cr)", bold=True, color=WHITE, fill=BLUE, align="center")
put(M, "E9", "1 Cr = ₹10,000,000", italic=True, size=9, color=WHITE, fill=BLUE, align="center")
tiers = [
    ("First 250",       0.000250, f"=MAX(0,MIN({aumCr},250))"),
    ("250 – 750",   0.000225, f"=MAX(0,MIN({aumCr},750)-250)"),
    ("750 – 1,500", 0.000175, f"=MAX(0,MIN({aumCr},1500)-750)"),
    ("1,500 – 2,500",0.000125, f"=MAX(0,MIN({aumCr},2500)-1500)"),
    ("Above 2,500",     0.000085, f"=MAX(0,{aumCr}-2500)"),
]
r = 10
for label, rate, aum_in in tiers:
    put(M, f"A{r}", label, align="left")
    put(M, f"B{r}", rate, align="center", fmt=PCT4)
    put(M, f"D{r}", aum_in, align="right", fmt=NUM2)
    put(M, f"C{r}", f"=D{r}*10000000*B{r}", align="right", fmt=INR)
    put(M, f"E{r}", "", border=True)
    r += 1
put(M, "A15", "Computed monthly SaaS", bold=True, align="left", fill=LIGHTBLUE)
put(M, "B15", "", fill=LIGHTBLUE)
put(M, "C15", "=SUM(C10:C14)", bold=True, align="right", fmt=INR, fill=LIGHTBLUE)
put(M, "D15", "=SUM(D10:D14)", align="right", fmt=NUM2, fill=LIGHTBLUE)
put(M, "E15", "Σ slab fees  (total AUM in col D)", italic=True, size=9, align="left", fill=LIGHTBLUE)
put(M, "A16", "Monthly minimum floor", align="left")
put(M, "B16", "per month", align="center", size=9)
put(M, "C16", f"={m2pFloor}", align="right", fmt=INR)
put(M, "D16", "", border=True)
put(M, "E16", "₹4L multi-LSP / ₹1.5L Meesho-only.", italic=True, size=9, align="left", wrap=True)
put(M, "A17", "Monthly recurring fee (billed)", bold=True, align="left", fill=GREEN)
put(M, "B17", "per month", align="center", size=9, fill=GREEN)
put(M, "C17", "=MAX(C15,C16)", bold=True, align="right", fmt=INR, fill=GREEN)
put(M, "D17", f"=C17/{fx}", align="right", fmt=EUR, fill=GREEN)
put(M, "E17", "Higher of computed SaaS vs floor.", italic=True, size=9, align="left", wrap=True, fill=GREEN)

section(M, "A19:E19", "RECURRING FEES  (per year)")
mrow(20, "Annual SaaS / platform fee", "per year", "=12*C17", "= 12 × monthly recurring fee above.")
mrow(21, "Subtotal — recurring / year", "", "=C20", "", bold=True, fill=LIGHTBLUE)

section(M, "A23:E23", "VARIABLE / BILLED ON ACTUALS  (excluded unless entered)")
mrow(24, "Customization / new reports", "₹17,000 / man-day", f"=17000*{m2pCustMd}",
     "Anything beyond the standard pack. Year-1 estimate from man-days.")
put(M, "A25", "SMS / cloud telephony", align="left")
put(M, "B25", "on actuals", align="center", size=9)
put(M, "C25", "Pass-through", align="center", size=9, italic=True)
put(M, "D25", "Pass-through", align="center", size=9, italic=True)
put(M, "E25", "Extra if opted, as applicable.", italic=True, size=9, align="left", wrap=True)

section(M, "A27:E27", "TOTALS")
mrow(28, "YEAR 1  (one-time + recurring + cust.)", "Year 1", "=C6+C21+C24", "", bold=True, fill=GREEN)
mrow(29, "Year 2  (recurring)", "Year 2", "=C21", "", fill=GREY)
mrow(30, "Year 3  (recurring)", "Year 3", "=C21", "", fill=GREY)
mrow(31, "3-YEAR TCO", "= horizon", f"=C6+C21*{horizon}+C24", "Incl. one-time only if toggle=1. 36-mo lock-in.", bold=True, fill=GREEN)
M.freeze_panes = "A4"

M2P = "'Vendor B - M2P'"
m_one_inr, m_one_eur = f"{M2P}!$C$6", f"{M2P}!$D$6"
m_mon_inr, m_mon_eur = f"{M2P}!$C$17", f"{M2P}!$D$17"
m_rec_inr, m_rec_eur = f"{M2P}!$C$21", f"{M2P}!$D$21"
m_y1_inr,  m_y1_eur  = f"{M2P}!$C$28", f"{M2P}!$D$28"
m_tco_inr, m_tco_eur = f"{M2P}!$C$31", f"{M2P}!$D$31"

# =====================================================================================
# SHEET 4: COMPARISON  (template-style: INR block + EUR block)
# =====================================================================================
C = wb.create_sheet("Comparison")
C.sheet_properties.tabColor = "70AD47"
cw = {"A": 38, "B": 22, "C": 20, "D": 20, "E": 16, "F": 2, "G": 1}
for c, w in cw.items():
    C.column_dimensions[c].width = w

title_row(C, "A1:E1", "VENDOR PRICING COMPARISON  —  Yubi (A) vs M2P (B)  ·  €100M/yr disbursal")

# verdict box
section(C, "A3:E3", "HEADLINE  (recalculates with your assumptions)")
put(C, "A4", "Lower 3-year TCO", bold=True, align="left", fill=YELLOW)
put(C, "B4", f'=IF({y_tco_inr}<{m_tco_inr},"Yubi (Vendor A)","M2P (Vendor B)")',
    bold=True, align="center", fill=YELLOW)
put(C, "C4", "3-yr saving (₹)", bold=True, align="center", fill=YELLOW)
put(C, "D4", f"=ABS({y_tco_inr}-{m_tco_inr})", bold=True, align="right", fmt=INR, fill=YELLOW)
put(C, "E4", f"=ABS({y_tco_eur}-{m_tco_eur})", bold=True, align="right", fmt=EUR, fill=YELLOW)
put(C, "A5", "Lower Year-1 cost", bold=True, align="left")
put(C, "B5", f'=IF({y_y1_inr}<{m_y1_inr},"Yubi (Vendor A)","M2P (Vendor B)")', bold=True, align="center")
put(C, "C5", "Yr-1 saving (₹)", bold=True, align="center")
put(C, "D5", f"=ABS({y_y1_inr}-{m_y1_inr})", bold=True, align="right", fmt=INR)
put(C, "E5", f"=ABS({y_y1_eur}-{m_y1_eur})", bold=True, align="right", fmt=EUR)
put(C, "A6", "Key driver", italic=True, align="left", size=9, border=False)
C.merge_cells("B6:E6")
put(C, "B6", "Yubi recurring = 0.20% of disbursal (flat, tenor-independent). M2P recurring scales with AUM/tenor. See Sensitivity tab.",
    italic=True, align="left", size=9, wrap=True, border=False)

def cmp_block(start, cur, fmt, y, m):
    """Render one comparison block (INR or EUR). y/m are dicts of formula refs."""
    section(C, f"A{start}:E{start}", f"ALL FIGURES IN {cur}")
    h = start + 1
    for col, txt in zip("ABCDE", ["Line item", "Frequency", "Yubi (A)", "M2P (B)", "Lower"]):
        put(C, f"{col}{h}", txt, bold=True, color=WHITE, fill=NAVY, align="center")
    rows = [
        ("Setup / one-time fees", "one-time", y["one"], m["one"], True),
        ("Fixed annual fee (AMC)", "per year", y["amc"], m["amc"], False),
        ("Variable platform fee", "per month", y["mon"], m["mon"], False),
        ("Variable platform fee", "per year", y["var"], m["var"], False),
        ("Recurring total / year", "per year", y["rec"], m["rec"], True),
    ]
    rr = h + 1
    for item, freq, yv, mv, bold in rows:
        fill = LIGHTBLUE if bold else None
        put(C, f"A{rr}", item, bold=bold, align="left", fill=fill)
        put(C, f"B{rr}", freq, align="center", size=9, fill=fill)
        put(C, f"C{rr}", yv, align="right", fmt=fmt, bold=bold, fill=fill)
        put(C, f"D{rr}", mv, align="right", fmt=fmt, bold=bold, fill=fill)
        put(C, f"E{rr}", f'=IF(C{rr}<D{rr},"A",IF(C{rr}>D{rr},"B","="))', align="center", fill=fill)
        rr += 1
    # excluded note
    put(C, f"A{rr}", "Billed on actuals (excluded)", italic=True, align="left", size=9)
    C.merge_cells(f"B{rr}:E{rr}")
    put(C, f"B{rr}", "Yubi: AA/3rd-party APIs + change requests @₹20k/man-day. M2P: SMS/telephony + customization @₹17k/man-day.",
        italic=True, align="left", size=9, wrap=True)
    rr += 1
    # totals
    put(C, f"A{rr}", "YEAR 1 TOTAL INVESTMENT", bold=True, align="left", fill=GREEN)
    put(C, f"B{rr}", "Year 1", align="center", size=9, fill=GREEN)
    put(C, f"C{rr}", y["y1"], align="right", fmt=fmt, bold=True, fill=GREEN)
    put(C, f"D{rr}", m["y1"], align="right", fmt=fmt, bold=True, fill=GREEN)
    put(C, f"E{rr}", f'=IF(C{rr}<D{rr},"A",IF(C{rr}>D{rr},"B","="))', align="center", bold=True, fill=GREEN)
    rr += 1
    put(C, f"A{rr}", "3-YEAR TCO", bold=True, align="left", fill=GREEN)
    put(C, f"B{rr}", "horizon", align="center", size=9, fill=GREEN)
    put(C, f"C{rr}", y["tco"], align="right", fmt=fmt, bold=True, fill=GREEN)
    put(C, f"D{rr}", m["tco"], align="right", fmt=fmt, bold=True, fill=GREEN)
    put(C, f"E{rr}", f'=IF(C{rr}<D{rr},"A",IF(C{rr}>D{rr},"B","="))', align="center", bold=True, fill=GREEN)
    return rr

# Yubi monthly variable (computed inline); M2P fixed AMC = none
y_inr = {"one": f"={y_one_inr}", "amc": f"={y_amc_inr}", "mon": f"=0.002*{mDisbINR}",
         "var": f"={y_var_inr}", "rec": f"={y_rec_inr}", "y1": f"={y_y1_inr}", "tco": f"={y_tco_inr}"}
m_inr = {"one": f"={m_one_inr}", "amc": 0, "mon": f"={m_mon_inr}",
         "var": f"={m_rec_inr}", "rec": f"={m_rec_inr}", "y1": f"={m_y1_inr}", "tco": f"={m_tco_inr}"}
end_inr = cmp_block(8, "INR  (₹)", INR, y_inr, m_inr)

y_eur = {"one": f"={y_one_eur}", "amc": f"={y_amc_eur}", "mon": f"=0.002*{mDisbEUR}",
         "var": f"={y_var_eur}", "rec": f"={y_rec_eur}", "y1": f"={y_y1_eur}", "tco": f"={y_tco_eur}"}
m_eur = {"one": f"={m_one_eur}", "amc": 0, "mon": f"={m_mon_eur}",
         "var": f"={m_rec_eur}", "rec": f"={m_rec_eur}", "y1": f"={m_y1_eur}", "tco": f"={m_tco_eur}"}
cmp_block(end_inr + 2, "EUR  (€)", EUR, y_eur, m_eur)
C.freeze_panes = "A2"

# =====================================================================================
# SHEET 5: SENSITIVITY
# =====================================================================================
S = wb.create_sheet("Sensitivity")
S.sheet_properties.tabColor = "BF9000"
for c, w in {"A": 30, "B": 22, "C": 22, "D": 22, "E": 22}.items():
    S.column_dimensions[c].width = w

title_row(S, "A1:E1", "SENSITIVITY  —  recurring cost vs loan tenor & book amortisation (€/yr)")
put(S, "A2", "Yubi recurring is flat (0.20% of disbursal, tenor-independent). M2P recurring scales with AUM.",
    italic=True, size=9, border=False)

put(S, "A4", "Avg tenor (months)", bold=True, color=WHITE, fill=NAVY, align="center")
put(S, "B4", "M2P — bullet (factor 1.0)", bold=True, color=WHITE, fill=NAVY, align="center")
put(S, "C4", "M2P — amortising (factor 0.55)", bold=True, color=WHITE, fill=NAVY, align="center")
put(S, "D4", "Yubi (flat)", bold=True, color=WHITE, fill=NAVY, align="center")
put(S, "E4", "Lower (amort. case)", bold=True, color=WHITE, fill=NAVY, align="center")

def m2p_annual_eur(tenor_m, fac):
    # AUM in Cr for this scenario
    aum = f"({disbINR}*({tenor_m}/12)*{fac}/10000000)"
    saas = (f"(MAX(0,MIN({aum},250))*10000000*0.000250"
            f"+MAX(0,MIN({aum},750)-250)*10000000*0.000225"
            f"+MAX(0,MIN({aum},1500)-750)*10000000*0.000175"
            f"+MAX(0,MIN({aum},2500)-1500)*10000000*0.000125"
            f"+MAX(0,{aum}-2500)*10000000*0.000085)")
    return f"=12*MAX({saas},{m2pFloor})/{fx}"

r = 5
for t in (3, 6, 9, 12):
    put(S, f"A{r}", t, align="center", bold=True)
    put(S, f"B{r}", m2p_annual_eur(t, "1.0"), align="right", fmt=EUR)
    put(S, f"C{r}", m2p_annual_eur(t, "0.55"), align="right", fmt=EUR)
    put(S, f"D{r}", f"=0.002*{disbEUR}", align="right", fmt=EUR)
    put(S, f"E{r}", f'=IF(C{r}<D{r},"M2P","Yubi")', align="center")
    r += 1

put(S, f"A{r+1}", "Reads: at the agreed 3–6 mo tenor M2P's AUM-based recurring sits well below Yubi's flat €200k/yr; "
                  "the gap narrows as tenor (and book) grows.", italic=True, size=9, border=False)
S.merge_cells(f"A{r+1}:E{r+1}")

# =====================================================================================
# SHEET 6: SCORECARD  (qualitative, 2 vendors, editable)
# =====================================================================================
SC = wb.create_sheet("Scorecard (optional)")
SC.sheet_properties.tabColor = "808080"
for c, w in {"A": 40, "B": 12, "C": 12, "D": 12, "E": 54}.items():
    SC.column_dimensions[c].width = w

title_row(SC, "A1:E1", "QUALITATIVE SCORECARD  (optional)  —  fill scores 1–5 in the shaded cells")
put(SC, "A2", "Pricing lives in the Comparison tab. This mirrors the scorecard template's weighted rollup for a fuller view.",
    italic=True, size=9, border=False)

cats = [
    ("1. Adherence to RFP Instructions", 0.05),
    ("2. Company Information", 0.05),
    ("3. Project Understanding", 0.20),
    ("4. Requirements", 0.30),
    ("5. Product Viability & History", 0.05),
    ("6. Terms & Conditions", 0.05),
    ("7. Vendor Software Demonstration", 0.25),
    ("8. Fee Summary", 0.05),
]
put(SC, "A4", "Criteria", bold=True, color=WHITE, fill=NAVY, align="left")
put(SC, "B4", "Weight", bold=True, color=WHITE, fill=NAVY, align="center")
put(SC, "C4", "Yubi (A)", bold=True, color=WHITE, fill=NAVY, align="center")
put(SC, "D4", "M2P (B)", bold=True, color=WHITE, fill=NAVY, align="center")
put(SC, "E4", "Notes / basis for score", bold=True, color=WHITE, fill=NAVY, align="left")
r = 5
for name, wt in cats:
    put(SC, f"A{r}", name, align="left")
    put(SC, f"B{r}", wt, align="center", fmt=NUM2, fill=INPUTFILL)
    put(SC, f"C{r}", 3, align="center", fill=INPUTFILL)
    put(SC, f"D{r}", 3, align="center", fill=INPUTFILL)
    put(SC, f"E{r}", "", align="left")
    r += 1
put(SC, f"A{r}", "Weight check (must = 1.00)", bold=True, align="left", fill=GREY)
put(SC, f"B{r}", f"=SUM(B5:B{r-1})", bold=True, align="center", fmt=NUM2, fill=GREY)
put(SC, f"C{r}", "", fill=GREY); put(SC, f"D{r}", "", fill=GREY); put(SC, f"E{r}", "", fill=GREY)
r += 1
put(SC, f"A{r}", "WEIGHTED TOTAL SCORE", bold=True, align="left", fill=GREEN)
put(SC, f"B{r}", "", fill=GREEN)
put(SC, f"C{r}", f"=SUMPRODUCT(B5:B{r-3},C5:C{r-3})", bold=True, align="center", fmt=NUM2, fill=GREEN)
put(SC, f"D{r}", f"=SUMPRODUCT(B5:B{r-3},D5:D{r-3})", bold=True, align="center", fmt=NUM2, fill=GREEN)
put(SC, f"E{r}", "Higher = better.", italic=True, size=9, align="left", fill=GREEN)

wb.active = wb["Comparison"]
wb.save(OUT)
print("WROTE", OUT)
