"""Shared helpers for E2E edge coverage scripts."""
from __future__ import annotations

import json
import shutil
import sys
import time
import uuid
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parent.parent.parent
MATRIX = ROOT / "e2e-test-matrix.xlsx"
RUNS = ROOT / ".e2e-runs"
COL_STATUS, COL_ACTUAL, COL_STEPS, COL_NOTES = 4, 5, 11, 13


def resolve_npx() -> str:
    """Return an npx executable that works on Windows (npx.cmd)."""
    candidates = ("npx.cmd", "npx.exe", "npx") if sys.platform == "win32" else ("npx",)
    for name in candidates:
        path = shutil.which(name)
        if path:
            return path
    return "npx"


def load_config() -> dict[str, str]:
    cfg = {
        "BASE_URL": "http://localhost:8080",
        "FRONTEND_URL": "http://localhost:5173",
        "ADMIN_USERNAME": "ops.admin",
        "ADMIN_PASSWORD": "ChangeMe123!",
        "E2E_FIXTURE_PREFIX": "E2E-EDGE",
        "E2E_EC068_MIN_LOANS": "100",
        "E2E_EC068_AUTO_SEED": "true",
    }
    env_file = Path(__file__).parent / "config.env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                cfg[k.strip()] = v.strip()
    return cfg


def wait_for_backend(cfg: dict, *, timeout_sec: float = 120, interval_sec: float = 2) -> None:
    """Probe login — avoids /actuator/health which blocks ~30s when the DB pool is exhausted."""
    deadline = time.time() + timeout_sec
    last_err: Exception | None = None
    url = f"{cfg['BASE_URL']}/api/v1/auth/login"
    body = {"username": cfg["ADMIN_USERNAME"], "password": cfg["ADMIN_PASSWORD"]}
    while time.time() < deadline:
        try:
            r = requests.post(url, json=body, timeout=5)
            if r.status_code == 200 and r.json().get("accessToken"):
                return
            last_err = RuntimeError(f"login returned {r.status_code}")
        except Exception as exc:
            last_err = exc
        time.sleep(interval_sec)
    raise TimeoutError(f"Backend not ready at {cfg['BASE_URL']}: {last_err}")


def login_password(cfg: dict, username: str, password: str) -> tuple[int, dict]:
    """POST /auth/login with 429 backoff (rate-limit aware)."""
    url = f"{cfg['BASE_URL']}/api/v1/auth/login"
    body = {"username": username, "password": password}
    last: requests.Response | None = None
    for _ in range(8):
        last = requests.post(url, json=body, timeout=30)
        if last.status_code == 429:
            retry_after = int(last.headers.get("Retry-After", "15") or 15)
            time.sleep(min(max(retry_after, 5), 60))
            continue
        return last.status_code, last.json() if last.content else {}
    if last is not None:
        return last.status_code, last.json() if last.content else {}
    return 0, {}


def admin_token(cfg: dict) -> str:
    wait_for_backend(cfg)
    code, body = login_password(cfg, cfg["ADMIN_USERNAME"], cfg["ADMIN_PASSWORD"])
    if code != 200 or not body.get("accessToken"):
        raise requests.HTTPError(f"admin login failed → {code}", response=None)
    return body["accessToken"]


def refresh_lsp_client_secret(cfg: dict, admin_bearer: str, client_id: str) -> str:
    """Rotate API client secret when fixture JSON is stale (e.g. after EC-008)."""
    base = cfg["BASE_URL"]
    ah = {"Authorization": f"Bearer {admin_bearer}"}
    clients = requests.get(f"{base}/api/v1/internal/admin/api-clients", headers=ah, timeout=30)
    clients.raise_for_status()
    row = next((c for c in clients.json() if c.get("clientId") == client_id), None)
    if not row or not row.get("id"):
        raise RuntimeError(f"API client not found for clientId={client_id}")
    rot = requests.post(
        f"{base}/api/v1/internal/admin/api-clients/{row['id']}/rotate-secret",
        headers=ah,
        timeout=30,
    )
    rot.raise_for_status()
    return rot.json()["clientSecret"]


def lsp_token_for_fixture(cfg: dict, fixture: dict, *, admin_bearer: str | None = None) -> str:
    """Obtain LSP token; refresh fixture clientSecret on 401; backoff on 429."""
    base = cfg["BASE_URL"]
    client_id = fixture["clientId"]
    secret = fixture["clientSecret"]
    url = f"{base}/api/v1/auth/token"
    body = {"clientId": client_id, "clientSecret": secret}
    r: requests.Response | None = None
    for _ in range(8):
        r = requests.post(url, json=body, timeout=30)
        if r.status_code == 429:
            retry_after = int(r.headers.get("Retry-After", "15") or 15)
            time.sleep(min(max(retry_after, 5), 60))
            continue
        break
    assert r is not None
    if r.status_code == 401:
        secret = refresh_lsp_client_secret(cfg, admin_bearer or admin_token(cfg), client_id)
        fixture["clientSecret"] = secret
        body = {"clientId": client_id, "clientSecret": secret}
        for _ in range(8):
            r = requests.post(url, json=body, timeout=30)
            if r.status_code == 429:
                retry_after = int(r.headers.get("Retry-After", "15") or 15)
                time.sleep(min(max(retry_after, 5), 60))
                continue
            break
    r.raise_for_status()
    return r.json()["accessToken"]


def update_matrix(results: list[dict]) -> int:
    wb = openpyxl.load_workbook(MATRIX)
    n = 0
    for item in results:
        tid = item["id"]
        for sn in ("Use Cases", "Edge Cases"):
            ws = wb[sn]
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == tid:
                    ws.cell(r, COL_STATUS, item["status"])
                    ws.cell(r, COL_ACTUAL, item["actual"])
                    if item.get("steps"):
                        ws.cell(r, COL_STEPS, item["steps"])
                    if item.get("notes"):
                        ws.cell(r, COL_NOTES, item["notes"])
                    n += 1
                    break
    wb.save(MATRIX)
    return n


def result(tid: str, status: str, actual: str, steps: str = "", notes: str = "") -> dict:
    return {"id": tid, "status": status, "actual": actual, "steps": steps, "notes": notes}


def suffix() -> str:
    return f"{int(time.time())}{uuid.uuid4().hex[:6]}"


def uuid_v4() -> str:
    return str(uuid.uuid4())


def _normalize_installment(inst: dict) -> dict:
    raw_amount = inst.get("amount", inst.get("installmentAmount"))
    return {
        "id": inst["id"],
        "amount": float(raw_amount) if raw_amount is not None else 0.0,
        "n": inst.get("n", inst.get("installmentNumber")),
        "status": inst.get("status"),
    }


def fetch_unpaid_installments(ah: dict, base: str, app_id: str) -> list[dict]:
    """Return schedule rows that can still accept a payment."""
    r = requests.get(
        f"{base}/api/v1/internal/ops/loan-applications/{app_id}/repayment-schedule",
        headers=ah,
        timeout=30,
    )
    if not r.ok:
        return []
    body = r.json()
    installments = body if isinstance(body, list) else (body.get("installments") or [])
    return [
        _normalize_installment(inst)
        for inst in installments
        if inst.get("status") not in ("PAID", "SETTLED", "CLOSED", "WAIVED")
    ]


def write_results(path: Path, results: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(results, indent=2), encoding="utf-8")


_ROLE_PRIORITY = (
    "SYSTEM_ADMIN",
    "OPS_USER",
    "PRODUCT_ADMIN",
    "LSP_UI_WRITE",
    "LSP_UI_READ",
    "LSP_API_CLIENT",
)


def build_playwright_storage_state(cfg: dict, path: Path) -> None:
    """One API login → Playwright storage state with bhawana-lms-session in localStorage."""
    from datetime import datetime, timedelta, timezone

    base = cfg["BASE_URL"].rstrip("/")
    frontend = cfg.get("FRONTEND_URL", "http://localhost:5173").rstrip("/")
    http = requests.Session()
    login = http.post(
        f"{base}/api/v1/auth/login",
        json={"username": cfg["ADMIN_USERNAME"], "password": cfg["ADMIN_PASSWORD"]},
        timeout=30,
    )
    login.raise_for_status()
    token_body = login.json()
    access = token_body["accessToken"]
    expires_in = max(60, int(token_body.get("expiresInSeconds") or 3600))
    expires_at = (
        datetime.now(timezone.utc) + timedelta(seconds=expires_in)
    ).isoformat().replace("+00:00", "Z")

    ctx = requests.get(
        f"{base}/api/v1/internal/system/context",
        headers={"Authorization": f"Bearer {access}"},
        timeout=30,
    )
    ctx.raise_for_status()
    context = ctx.json()
    roles = context.get("roles") or []
    role = next((r for r in _ROLE_PRIORITY if r in roles), "OPS_USER")
    session = {
        "user": {
            "id": context["id"],
            "username": context["username"],
            "role": role,
            "lspId": context.get("lspId"),
            "mustChangePassword": bool(token_body.get("passwordChangeRequired")),
        },
        "accessToken": access,
        "expiresAt": expires_at,
    }
    cookies = []
    for cookie in http.cookies:
        cookies.append(
            {
                "name": cookie.name,
                "value": cookie.value,
                "domain": "localhost",
                "path": cookie.path or "/api/v1/auth",
                "httpOnly": True,
                "secure": False,
                "sameSite": "Strict",
            }
        )

    state = {
        "cookies": cookies,
        "origins": [
            {
                "origin": frontend,
                "localStorage": [
                    {"name": "bhawana-lms-session", "value": json.dumps(session)},
                ],
            }
        ],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state), encoding="utf-8")
