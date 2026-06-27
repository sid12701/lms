"""
Generates D:/Desktop-New/Folders/LMS/e2e-test-matrix.xlsx.

Two sheets — `Use Cases` (PDF UC-001..UC-047 + repo-derived additions) and
`Edge Cases` (negative paths, validation, state-machine, RBAC, integration).

Required columns (exact order, both sheets):
  UC | The use case | Expected behaviour | Pass / Fail | Actual behaviour / Current behaviour

Additional columns (after the required ones):
  Actor | Module | Source | Test Type | API / Screen | Steps Performed | Severity | Notes

All Pass/Fail values are `Not Tested` per user instruction (audit-only run);
`Actual behaviour` is derived from code reading, not live execution.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUT_PATH = r"D:\Desktop-New\Folders\LMS\e2e-test-matrix.xlsx"

COLUMNS = [
    ("UC", 10),
    ("The use case", 55),
    ("Expected behaviour", 70),
    ("Pass / Fail", 14),
    ("Actual behaviour / Current behaviour", 60),
    ("Actor", 18),
    ("Module", 22),
    ("Source", 22),
    ("Test Type", 14),
    ("API / Screen", 50),
    ("Steps Performed", 50),
    ("Severity", 12),
    ("Notes", 50),
]

PASS_FAIL_DEFAULT = "Not Tested"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F6FB")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_sheet(ws: Worksheet, row_count: int) -> None:
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 32

    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = WRAP
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
        ws.row_dimensions[row_idx].height = 75

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row_count + 1}"


def write_rows(ws: Worksheet, rows: list[dict]) -> None:
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=row["id"])
        ws.cell(row=r_idx, column=2, value=row["use_case"])
        ws.cell(row=r_idx, column=3, value=row["expected"])
        ws.cell(row=r_idx, column=4, value=row.get("pass_fail", PASS_FAIL_DEFAULT))
        ws.cell(row=r_idx, column=5, value=row.get("actual", "Not yet executed — derived from code; awaits live run."))
        ws.cell(row=r_idx, column=6, value=row["actor"])
        ws.cell(row=r_idx, column=7, value=row["module"])
        ws.cell(row=r_idx, column=8, value=row["source"])
        ws.cell(row=r_idx, column=9, value=row["test_type"])
        ws.cell(row=r_idx, column=10, value=row["api_or_screen"])
        ws.cell(row=r_idx, column=11, value=row.get("steps", "Awaiting live execution."))
        ws.cell(row=r_idx, column=12, value=row.get("severity", ""))
        ws.cell(row=r_idx, column=13, value=row.get("notes", ""))


def use_case_rows() -> list[dict]:
    """47 PDF use cases + a handful discovered in the codebase but not in the PDF."""
    rows: list[dict] = []

    def uc(num: int, title, expected, actor, module, api, *, source="PDF + code", steps="Awaiting live execution.", notes=""):
        rows.append({
            "id": f"UC-{num:03d}",
            "use_case": title,
            "expected": expected,
            "actor": actor,
            "module": module,
            "source": source,
            "test_type": "UI + API" if "/" in (api or "") and "screen" in (api or "").lower() else "API",
            "api_or_screen": api,
            "steps": steps,
            "severity": "",
            "notes": notes,
        })

    # ---------- AUTHENTICATION & ACCESS ----------
    uc(1, "User Login",
       "POST /api/v1/auth/login with valid credentials returns access + refresh tokens, role claims, and routes user to role-appropriate landing page (SYSTEM_ADMIN→/home, OPS_USER→/loan-applications, PRODUCT_ADMIN→/products, LSP→/my-loans).",
       "All human users", "Authentication", "POST /api/v1/auth/login + /login UI",
       steps="1) Open /login. 2) Submit ops.admin / ChangeMe123!. 3) Verify access token, role claim, landing redirect, session badge.")

    uc(2, "Mandatory Password Change",
       "After login when must_change_password=true, user is redirected to /change-password; POST /api/v1/auth/password updates password, clears flag, allows normal access.",
       "Authenticated user", "Authentication", "POST /api/v1/auth/password + /change-password UI",
       steps="1) Reset a user's password via UserAdmin (sets must_change_password=true). 2) Log in as that user. 3) Confirm forced redirect; submit new password.")

    uc(3, "Session Refresh and Logout",
       "POST /auth/refresh renews access token without re-credentialing; POST /auth/logout invalidates refresh token and clears session.",
       "All human users", "Authentication", "POST /api/v1/auth/refresh, POST /api/v1/auth/logout",
       steps="1) Login. 2) Call /refresh near expiry. 3) Call /logout and confirm refresh token rejected on next refresh attempt.")

    uc(4, "API Client Token Issuance",
       "POST /api/v1/auth/token with valid client_id+secret returns scoped JWT (tenant=lsp_id, role=LSP_API_CLIENT). Inactive client, inactive LSP, or IP allowlist violation returns 401/403.",
       "LSP API Client", "Authentication", "POST /api/v1/auth/token",
       steps="1) Create API client (UC-012). 2) POST credentials. 3) Decode JWT and confirm scope.")

    uc(5, "View Session Context",
       "GET /api/v1/internal/system/context returns {userId, username, roles[], lspId?, tenantScope} for authenticated user.",
       "All authenticated humans", "Session", "GET /api/v1/internal/system/context",
       steps="1) Authenticate. 2) Hit /context. 3) Verify tenant scope matches expected.")

    # ---------- PLATFORM & TENANT ADMIN ----------
    uc(6, "Create LSP Tenant",
       "POST /api/v1/internal/admin/lsps (SYSTEM_ADMIN) creates inactive LSP record; new id is then available for product mappings, users, API clients, webhook config.",
       "System Administrator", "LSP Admin", "POST /api/v1/internal/admin/lsps + /lsps UI",
       steps="1) Login as SYSTEM_ADMIN. 2) Open /lsps. 3) Create LSP. 4) Verify it appears in list and audit event recorded.")

    uc(7, "Activate or Deactivate LSP",
       "PUT /lsps/{lspId}/status flips LSP status; deactivation cascades to API clients (disabled) and blocks LSP API/UI access.",
       "System Administrator", "LSP Admin", "PUT /api/v1/internal/admin/lsps/{lspId}/status",
       steps="1) Deactivate LSP. 2) Attempt LSP API call → expect 401/403. 3) Reactivate; confirm restored.")

    uc(8, "Configure LSP Webhook Subscription",
       "PUT /lsps/{lspId}/webhook-subscription persists endpoint URL + secret + event subscriptions; future domain events for that LSP enqueue in webhook_event_outbox.",
       "System Administrator", "Webhooks", "PUT /api/v1/internal/admin/lsps/{lspId}/webhook-subscription")

    uc(9, "Manage LSP IP Allowlists",
       "PUT/DELETE on /lsps/{lspId}/api-ip-allowlist and /ui-ip-allowlist plus enforcement toggle controls per-surface IP restrictions; enforcement applied via LspSurfaceIpAllowlistFilter.",
       "System Administrator", "Security/IP Allowlist", "/api/v1/internal/admin/lsps/{lspId}/api-ip-allowlist, /ui-ip-allowlist, /allowlist-enforcement")

    # ---------- USERS & API CLIENTS ----------
    uc(10, "Create Internal or LSP User",
        "POST /api/v1/internal/admin/users (SYSTEM_ADMIN) creates user with role(s) and optional LSP scope; must_change_password flag set on initial password.",
        "System Administrator", "User Admin", "POST /api/v1/internal/admin/users + /users UI",
        steps="1) Open /users. 2) Create LSP_UI_WRITE, LSP_UI_READ, OPS_USER, PRODUCT_ADMIN, LSP_API_CLIENT users. 3) Confirm role and tenant scoping.")

    uc(11, "Update User or Reset Password",
        "PUT /users/{userId} edits status/roles (with last-SYSTEM_ADMIN + self-disable guards). POST /users/{userId}/reset-password mints temporary password and forces change on next login. POST /users/{userId}/revoke-sessions bumps token_version invalidating active JWTs.",
        "System Administrator", "User Admin", "PUT /api/v1/internal/admin/users/{userId}, POST /reset-password, POST /revoke-sessions")

    uc(12, "Create API Client",
        "POST /api/v1/internal/admin/api-clients (SYSTEM_ADMIN) creates client tied to LSP; returns client_id and one-time secret. Audit event recorded.",
        "System Administrator", "API Clients", "POST /api/v1/internal/admin/api-clients + /api-clients UI")

    uc(13, "Rotate API Client Secret",
        "POST /api-clients/{id}/rotate-secret issues new secret; grace period (default 300s) keeps old secret valid for cutover; rotation audited.",
        "System Administrator", "API Clients", "POST /api/v1/internal/admin/api-clients/{id}/rotate-secret")

    # ---------- PRODUCT CATALOG ----------
    uc(14, "Create or Update Loan Product",
        "PUT /products/{productId} (SYSTEM_ADMIN/PRODUCT_ADMIN) defines product name, min/max principal, min/max tenure, interest rate, processing fee, status.",
        "System Administrator", "Products", "PUT /api/v1/internal/admin/products/{productId} + /products UI")

    uc(15, "Configure Product–LSP Mappings",
        "PUT /products/{productId}/mappings and POST /product-lsp-mappings/entries enable/disable per-LSP mapping; only enabled+ACTIVE mappings allow origination & appear in LSP catalog.",
        "System Administrator", "Products", "/api/v1/internal/admin/products/{productId}/mappings, /product-lsp-mappings/entries")

    # ---------- ORIGINATION ----------
    uc(16, "LSP API — Create Loan Application",
        "POST /api/v1/lsp/loan-applications (LSP_API_CLIENT) validates payload, dedups borrower by PAN, creates application + document checklist + intake snapshot, transitions to AWAITING_APPROVAL, fires auto-approval engine.",
        "LSP API Client", "LSP API / Origination", "POST /api/v1/lsp/loan-applications",
        steps="1) Obtain token (UC-004). 2) POST application with valid borrower+product+amount+tenure. 3) Confirm 201 + applicationId + AWAITING_APPROVAL→APPROVED_PENDING_DISBURSAL or REJECTED.")

    uc(17, "Operations — Create Loan Application",
        "Loan creation from operations console (BRD/PDF describes API-only today; CLAUDE.md memory ADR 0003 confirms LSP origination is API-only — no internal UI form). Service path exists via LoanApplicationService.createApplication.",
        "System Administrator / Operations User", "Origination", "Service: LoanApplicationService.createApplication (no UI form)",
        notes="ADR 0003: LSP origination is API-only by design. No internal UI create form exists. Treat as Blocked from UI side; API path is the LSP endpoint.")

    uc(18, "Auto-Approval Rule Evaluation",
        "LoanAutoApprovalRuleEngine evaluates 8 rules: PRODUCT_INACTIVE, LSP_INACTIVE, LSP_PRODUCT_MAPPING_INACTIVE, LOAN_AMOUNT_OUT_OF_RANGE, LOAN_TENURE_OUT_OF_RANGE, BORROWER_REQUIRED_FIELDS_MISSING, REQUIRED_DOCUMENTS_NOT_UPLOADED, BORROWER_HAS_OPEN_LOAN. All pass → APPROVED_PENDING_DISBURSAL + loan_account; any fail → REJECTED with rejection_reason_json.",
        "System (engine)", "Auto-Approval", "Service: LoanAutoApprovalRuleEngine.evaluate")

    uc(19, "Upload KYC Documents (LSP)",
        "POST /lsp/loan-applications/{id}/documents (JSON metadata, multipart single, or multipart batch) stores file via DocumentUploadPolicy and marks checklist item SUBMITTED. PDF/JPEG allow-list, max 10 MB (5 MB for PAN/AADHAAR).",
        "LSP API Client / LSP_UI_WRITE", "Documents", "POST /api/v1/lsp/loan-applications/{id}/documents{,/batch}")

    uc(20, "Download KYC Documents (Operations)",
        "GET /ops/loan-applications/{id}/kyc-documents/{type}/content streams file with Content-Disposition; access recorded in LoanApplicationDocumentAccessAudit. /download-all returns zip.",
        "System Administrator / Operations User", "Documents", "GET /api/v1/internal/ops/loan-applications/{id}/kyc-documents/...")

    uc(21, "Manual Status Transition",
        "POST /ops/loan-applications/{id}/status-transitions (SYSTEM_ADMIN) advances within state machine when canTransitionTo() allows; emits status-change webhook + audit.",
        "System Administrator", "Lifecycle", "POST /api/v1/internal/ops/loan-applications/{id}/status-transitions")

    uc(22, "Manual Status Override",
        "POST /ops/loan-applications/{id}/manual-status (SYSTEM_ADMIN) bypasses normal transition guards; reason code required (MANUAL_ADMIN_OVERRIDE); fully audited.",
        "System Administrator", "Lifecycle", "POST /api/v1/internal/ops/loan-applications/{id}/manual-status")

    uc(23, "Invalidate Loan (Pre-Disbursal)",
        "POST /lsp/loan-applications/{id}/invalid (LSP API or via /my-loans UI by LSP_UI_WRITE) requires Idempotency-Key + reason code (with optional text for OTHER); moves to terminal INVALID + cascades to loan_account; fires status-change webhook.",
        "LSP API Client / LSP_UI_WRITE", "Lifecycle", "POST /api/v1/lsp/loan-applications/{id}/invalid")

    uc(24, "Submit Repayment Schedule",
        "PUT /lsp/loan-applications/{id}/repayment-schedule validates and stores installment plan; schedule becomes a disbursement prerequisite and basis for payment allocation.",
        "LSP API Client", "Disbursement", "PUT /api/v1/lsp/loan-applications/{id}/repayment-schedule")

    uc(25, "Disbursement Bank Check",
        "POST /lsp/loan-applications/{id}/disbursement-bank-check validates borrower bank details vs disbursement requirements; mismatch logged in LoanDisbursementBankMismatchLog and surfaced as ops alert.",
        "LSP API Client", "Disbursement", "POST /api/v1/lsp/loan-applications/{id}/disbursement-bank-check")

    uc(26, "Initiate Disbursement",
        "POST /ops/loan-applications/{id}/disbursement-requests (SYSTEM_ADMIN) triggers LoanDisbursementService → adapter; updates loan account + application status per outcome (DISBURSED, DISBURSEMENT_RETRY).",
        "System Administrator", "Disbursement", "POST /api/v1/internal/ops/loan-applications/{id}/disbursement-requests")

    uc(27, "Automated Disbursement Processing",
        "LoanDisbursementWorker scheduled run finds APPROVED_PENDING_DISBURSAL / DISBURSEMENT_RETRY apps with met prerequisites, dispatches via adapter; success→DISBURSED→UNDER_REPAYMENT, retryable→DISBURSEMENT_RETRY (up to configured max), exhaustion→ops alert.",
        "System (worker)", "Disbursement", "Worker: LoanDisbursementWorker")

    uc(28, "Simulate Disbursement Outcome",
        "POST /ops/loan-applications/{id}/disbursement-requests/mock-outcome applies a chosen MockDisbursementOutcome (SUCCESS/RETRYABLE_FAILURE/PERMANENT_FAILURE); only active when mock adapter enabled.",
        "System Administrator", "Disbursement (Mock)", "POST /api/v1/internal/ops/loan-applications/{id}/disbursement-requests/mock-outcome")

    uc(29, "Record Payment (Internal Operations)",
        "POST /ops/loan-applications/{id}/payments (SYSTEM_ADMIN/OPS_USER) requires Idempotency-Key + targetInstallmentId + exact installment amount + channel (NEFT/RTGS/IMPS); allocates to that installment; advances DISBURSED→UNDER_REPAYMENT on first payment; CLOSED on last installment.",
        "System Administrator / Operations User", "Servicing", "POST /api/v1/internal/ops/loan-applications/{id}/payments")

    uc(30, "Record Payment (LSP API)",
        "POST /lsp/loans/{loanId}/payments mirrors ops payment validation (idempotent, full installment amount, channel); same status transitions and webhooks fire.",
        "LSP API Client", "Servicing", "POST /api/v1/lsp/loans/{loanId}/payments")

    uc(31, "Request Foreclosure Quote",
        "POST /lsp/loans/{loanId}/foreclosure-quote or /ops/.../foreclosure-quotes calculates payoff amount + validity period; quote stored for later execution.",
        "System Administrator / LSP API Client", "Servicing", "POST /api/v1/lsp/loans/{loanId}/foreclosure-quote, POST /api/v1/internal/ops/loan-applications/{id}/foreclosure-quotes")

    uc(32, "Execute Foreclosure",
        "POST /lsp/loans/{loanId}/foreclosure-quotes/{quoteId}/execute or ops equivalent applies foreclosure → terminal FORECLOSED; partner notified via webhook.",
        "System Administrator", "Servicing", "POST /api/v1/lsp/loans/{loanId}/foreclosure-quotes/{quoteId}/execute, POST /ops/loan-applications/{id}/foreclosure-quotes/{quoteId}/execute")

    uc(33, "Update Borrower Bank Details",
        "PATCH /lsp/borrowers/{borrowerId}/bank-details and ops equivalent persist bank account, ifsc, holder name; audited in BorrowerBankDetailsUpdateAudit; may trigger borrower.bank.updated webhook.",
        "System Administrator / LSP API Client", "Borrowers", "PATCH /api/v1/lsp/borrowers/{borrowerId}/bank-details, PATCH /api/v1/internal/admin/borrowers/{borrowerId}/bank-details")

    uc(34, "Search and View Borrowers",
        "GET /borrowers list (with filters) and GET /borrowers/{id} return borrower profile + linked loans + DPD aggregate; aadhaar masked in API responses.",
        "System Administrator / Operations User", "Borrowers", "GET /api/v1/internal/admin/borrowers, GET /borrowers/{id} + /borrowers UI")

    uc(35, "View Portfolio Dashboard",
        "GET /home/overview returns HomeDashboardSummary: totals (disbursed/active/overdue), applications-by-status, DPD buckets, openAlerts, avg approval TAT (30d), critical disbursements.",
        "System Administrator", "Home Dashboard", "GET /api/v1/internal/home/overview + /home UI")

    uc(36, "Search Loan Applications",
        "GET /ops/loan-applications supports filters (status, LSP, product, source, search, disbursal date range); cursor pagination; sensitive-data toggle masks PII columns.",
        "System Administrator / Operations User", "Loan Applications", "GET /api/v1/internal/ops/loan-applications + /loan-applications UI")

    uc(37, "Portfolio MIS — Preview and Sync Download",
        "GET /reports/portfolio-mis/preview returns first N rows (aadhaar + bank account masked). GET /reports/portfolio-mis (text/csv) streams full report for sync download. Date range + LSP filter.",
        "System Administrator", "Reports", "GET /api/v1/internal/reports/portfolio-mis/preview, /portfolio-mis (CSV) + /reports UI")

    uc(38, "Async MIS Report Generation",
        "POST /reports/portfolio-mis/requests queues report; ReportRequestProcessingWorker generates CSV in storage (R2 or local); GET /reports/requests + /requests/{id}/download retrieve.",
        "System Administrator", "Reports", "POST /api/v1/internal/reports/portfolio-mis/requests, GET /requests, GET /requests/{id}/download")

    uc(39, "Acknowledge Operations Alert",
        "POST /alerts/{alertId}/acknowledge with optional note (≤500 chars) marks alert ACKNOWLEDGED; visible in /alerts table for team coordination.",
        "System Administrator / Operations User", "Alerts", "POST /api/v1/internal/alerts/{alertId}/acknowledge + /alerts UI")

    uc(40, "Escalate Loan to Administrator",
        "POST /alerts/escalate (OPS_USER) creates OPS_USER_ESCALATION alert tied to loan application; ESCALATE button replaced Approve/Reject in OPS UI (Gap #16).",
        "Operations User", "Alerts", "POST /api/v1/internal/alerts/escalate + EscalateToAdminDialog UI")

    uc(41, "Scheduled Alert Rule Evaluation",
        "AlertRuleSchedulerWorker (5 min cadence) evaluates seven AlertRule entries (stale intake, stuck disbursement, DPD bucket transition, LSP reject spike, etc.); creates/updates OpsAlert deduped via createAlertIfAbsent.",
        "System (worker)", "Alerts", "Worker: AlertRuleSchedulerWorker, GET /api/v1/internal/alerts/rules")

    uc(42, "Webhook Outbox Dispatch",
        "WebhookOutboxDispatchWorker claims PENDING outbox rows, signs payload, POSTs to subscriber URL; PENDING→DELIVERED on 2xx, →RETRYABLE_FAILURE with backoff, →PERMANENT_FAILURE after max attempts.",
        "System (worker)", "Webhooks", "Worker: WebhookOutboxDispatchWorker, POST /api/v1/internal/admin/webhook-outbox/dispatch (manual fire)")

    uc(43, "Manual Webhook Redrive",
        "POST /webhook-outbox/{id}/redrive (SYSTEM_ADMIN) resets PERMANENT_FAILURE entry to PENDING; redrive audit recorded; subject to max redrive attempts.",
        "System Administrator", "Webhooks", "POST /api/v1/internal/admin/webhook-outbox/{id}/redrive")

    uc(44, "Audit Explorer Search",
        "GET /audit-events (SYSTEM_ADMIN) UNION ALL across APPLICATION/INTAKE/DOCUMENT_ACCESS/PRODUCT streams with filter pushdown (actorUsername, lspId, loanApplicationId, borrowerId, productId, since/until). Aadhaar masked in INTAKE detail. offset/limit (cap 500) + optional paginationDetails.",
        "System Administrator", "Audit", "GET /api/v1/internal/admin/audit-events + /audit UI")

    uc(45, "Auth Audit Search",
        "GET /api/v1/internal/ops/auth-audit lists AuthEventAudit (login success/failure, token issuance, password change) with filters. PDF flags 'no dedicated UI'; FE wires via AuditPage timeline streams.",
        "System Administrator", "Audit", "GET /api/v1/internal/ops/auth-audit",
        notes="PDF labelled 'Needs Clarification — no dedicated UI documented'. FE delivers it through /audit timeline; verify UI surfaces auth events.")

    uc(46, "LSP View Own Loans",
        "GET /lsp/loan-applications + /lsp/loan-applications/{id} + /lsp/loans/{id} return only the authenticated LSP's records via LspTenantContextInterceptor enforcement; /my-loans UI scoped to own tenant.",
        "LSP_UI_READ / LSP_UI_WRITE", "LSP Self-Service", "GET /api/v1/lsp/loan-applications, /lsp/loans/{id} + /my-loans UI")

    uc(47, "View LSP Product Catalog",
        "GET /api/v1/lsp/products returns only ACTIVE products with an enabled mapping for the authenticated LSP; mapping/product status changes reflected immediately.",
        "LSP API Client / LSP UI User", "Products / LSP", "GET /api/v1/lsp/products")

    # ---- Additions discovered in codebase, not in PDF ----
    uc(48, "Admin Metadata catalog (enums for FE pickers)",
        "GET /api/v1/internal/admin/metadata returns reason codes, status enums, document types, channels for FE dropdown population. Not in PDF.",
        "All internal roles", "Admin Metadata", "GET /api/v1/internal/admin/metadata",
        notes="Code-derived; not in PDF.")

    uc(49, "LSP Options for admin dropdowns",
        "GET /api/v1/internal/admin/lsp-options returns lightweight LSP list (id+name+status) for filter pickers in ops UIs. Not in PDF.",
        "All internal roles", "LSP Admin", "GET /api/v1/internal/admin/lsp-options",
        notes="Code-derived; not in PDF.")

    uc(50, "Per-application webhook events view (Gap #5)",
        "GET /ops/loan-applications/{id}/webhook-events returns ≤200 outbox rows newest-first for SYSTEM_ADMIN/OPS_USER with status mapped to PENDING/DELIVERED/FAILED/DEAD_LETTERED. Not in PDF.",
        "System Administrator / Operations User", "Webhooks", "GET /api/v1/internal/ops/loan-applications/{id}/webhook-events",
        notes="Code-derived per gap-fixes Gap #5; not in PDF.")

    uc(51, "Borrower bank details retrieval (LSP)",
        "GET /lsp/borrowers/{borrowerId}/bank-details returns current bank info for the authenticated LSP's borrower; PATCH update covered in UC-033.",
        "LSP API Client", "Borrowers", "GET /api/v1/lsp/borrowers/{borrowerId}/bank-details",
        notes="Code-derived; not explicitly in PDF.")

    uc(52, "Document access audit per application",
        "GET /ops/loan-applications/{id}/document-access-audits returns DOCUMENT_ACCESS audit rows for compliance review; access logged on every download.",
        "System Administrator / Operations User", "Audit", "GET /api/v1/internal/ops/loan-applications/{id}/document-access-audits",
        notes="Surfaced via Audit Explorer; ops endpoint exists.")

    return rows


def edge_case_rows() -> list[dict]:
    rows: list[dict] = []

    def ec(num: int, title, expected, actor, module, api, *, severity="Medium", source="Code audit", test_type="API", steps="Awaiting live execution.", notes=""):
        rows.append({
            "id": f"EC-{num:03d}",
            "use_case": title,
            "expected": expected,
            "actor": actor,
            "module": module,
            "source": source,
            "test_type": test_type,
            "api_or_screen": api,
            "steps": steps,
            "severity": severity,
            "notes": notes,
        })

    # ---------- AUTH / RBAC ----------
    ec(1, "Login with wrong password",
       "POST /auth/login with invalid password → 401 with no user enumeration; AuthEventAudit entry with failure reason BAD_CREDENTIALS.",
       "All humans", "Authentication", "POST /api/v1/auth/login",
       severity="High",
       steps="1) Submit valid username, wrong password. 2) Confirm 401 + no role disclosure. 3) Repeat to trigger brute-force lockout (#155 fix).")

    ec(2, "Login with unknown username",
       "POST /auth/login with non-existent user → 401 with identical body/timing as wrong-password to avoid user enumeration.",
       "All humans", "Authentication", "POST /api/v1/auth/login",
       severity="High")

    ec(3, "Brute-force lockout (issue #155)",
       "After N consecutive failed logins for same username, account is locked for cooldown window; alert raised; user receives lockout message on next attempt (resolved per PR #196).",
       "All humans", "Authentication", "POST /api/v1/auth/login",
       severity="High",
       notes="Closed via PR #196 (scheduler-driven lockout). Verify lockout window, alert generation, and unlock path.")

    ec(4, "Inactive user attempts login",
       "POST /auth/login for status=DISABLED user → 401/403 with disabled-account reason; AuthEventAudit captures.",
       "All humans", "Authentication", "POST /api/v1/auth/login",
       severity="High")

    ec(5, "Expired access token",
       "Internal endpoint with expired JWT → 401; FE silently refreshes via /auth/refresh; if refresh also expired, redirect to /login.",
       "All authenticated", "Authentication", "Any /api/v1/internal/* endpoint",
       severity="Medium")

    ec(6, "Tampered JWT (signature mismatch)",
       "Endpoint call with hand-edited JWT (changed sub or role) → 401; no role escalation possible.",
       "Attacker", "Authentication", "Any authenticated endpoint",
       severity="Critical")

    ec(7, "Token revoked via session invalidation",
       "After POST /users/{userId}/revoke-sessions or role change, prior access tokens fail with tv-claim mismatch → 401.",
       "System Administrator", "Authentication", "POST /api/v1/internal/admin/users/{userId}/revoke-sessions",
       severity="High",
       notes="token_version (tv) JWT claim per Gap #13.")

    ec(8, "API client with rotated secret (within grace)",
       "Old secret continues to work for grace window (300s default) after rotation; after expiry → 401.",
       "LSP API Client", "Authentication", "POST /api/v1/auth/token",
       severity="Medium")

    ec(9, "API client used by inactive LSP",
       "Valid client_id+secret on deactivated LSP → 401/403; LspStatusService cascade ensures client disabled.",
       "LSP API Client", "Authentication", "POST /api/v1/auth/token",
       severity="High")

    ec(10, "IP allowlist violation (LSP API)",
        "API call from IP not in API allowlist (enforcement enabled) → 403 with allowlist-violation reason; OpsAlert fires on repeated violations.",
        "LSP API Client", "Security/IP Allowlist", "Any /api/v1/lsp/* endpoint",
        severity="High")

    ec(11, "IP allowlist violation (LSP UI)",
        "/login or any /my-loans request from non-allowlisted IP for LSP UI user → 403.",
        "LSP UI User", "Security/IP Allowlist", "/login (LSP user), /my-loans",
        severity="High")

    ec(12, "Cross-tenant LSP access attempt",
        "LSP A's token used to GET /lsp/loan-applications/{id} where id belongs to LSP B → 404 (not 403, to avoid leaking existence).",
        "LSP API Client", "Tenancy / Isolation", "GET /api/v1/lsp/loan-applications/{id}",
        severity="Critical",
        notes="LspTenantContextInterceptor enforces; memory note re #89: TestTenantContextRestoreFilter is @Profile('test') only — risk of 401 loop on live backend with current admin filter chain.")

    ec(13, "Role-bypass attempt (LSP_API_CLIENT hitting admin endpoint)",
        "LSP_API_CLIENT JWT used against /api/v1/internal/admin/* → 403 via @PreAuthorize.",
        "LSP API Client", "Authorization", "Any /api/v1/internal/admin/* endpoint",
        severity="Critical")

    ec(14, "OPS_USER attempts SYSTEM_ADMIN-only endpoint",
        "OPS_USER calling /audit-events, /lsps, /products write, /api-clients, /reports → 403.",
        "Operations User", "Authorization", "Various SYSTEM_ADMIN-only endpoints",
        severity="High")

    ec(15, "PRODUCT_ADMIN attempts non-product endpoint",
        "PRODUCT_ADMIN trying /loan-applications/* or /lsps → 403.",
        "Product Admin", "Authorization", "/api/v1/internal/admin/lsps, /loan-applications",
        severity="High")

    ec(16, "Last SYSTEM_ADMIN cannot disable self",
        "PUT /users/{userId} disabling last active SYSTEM_ADMIN → 400 with last-admin guard message (Gap #13 guard).",
        "System Administrator", "User Admin", "PUT /api/v1/internal/admin/users/{userId}",
        severity="High")

    ec(17, "Self password reset by admin → forces change on next login",
        "POST /users/{userId}/reset-password sets must_change_password=true and revokes sessions; user must change on next login.",
        "System Administrator", "User Admin", "POST /api/v1/internal/admin/users/{userId}/reset-password",
        severity="Medium")

    # ---------- ORIGINATION VALIDATION (8 auto-approval rules) ----------
    ec(18, "Auto-rejection: PRODUCT_INACTIVE",
        "POST /lsp/loan-applications with productId whose status=INACTIVE → application created but auto-REJECTED with PRODUCT_INACTIVE in rejection_reason_json.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="High")

    ec(19, "Auto-rejection: LSP_INACTIVE",
        "Inactive LSP submits application → REJECTED with LSP_INACTIVE (though token issuance already fails at auth layer — UC-007).",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="High")

    ec(20, "Auto-rejection: LSP_PRODUCT_MAPPING_INACTIVE",
        "Product valid + LSP active, but mapping disabled → REJECTED with LSP_PRODUCT_MAPPING_INACTIVE.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="High")

    ec(21, "Auto-rejection: LOAN_AMOUNT_OUT_OF_RANGE (below min)",
        "loanAmount < product.minPrincipal → REJECTED with LOAN_AMOUNT_OUT_OF_RANGE.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    ec(22, "Auto-rejection: LOAN_AMOUNT_OUT_OF_RANGE (above max)",
        "loanAmount > product.maxPrincipal → REJECTED.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    ec(23, "Auto-rejection: LOAN_TENURE_OUT_OF_RANGE",
        "tenure < minTenureMonths or > maxTenureMonths → REJECTED with LOAN_TENURE_OUT_OF_RANGE.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    ec(24, "Auto-rejection: BORROWER_REQUIRED_FIELDS_MISSING",
        "Missing/blank any of: fullName, pan, mobile, aadharNumber, addressLine1, city, state, zip, monthlyIncome>0, referencePersonName, referencePersonNumber → REJECTED with BORROWER_REQUIRED_FIELDS_MISSING.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="High")

    ec(25, "Auto-rejection: REQUIRED_DOCUMENTS_NOT_UPLOADED",
        "Approval-required documents missing from checklist at AWAITING_APPROVAL → REJECTED with REQUIRED_DOCUMENTS_NOT_UPLOADED.",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="High")

    ec(26, "Auto-rejection: BORROWER_HAS_OPEN_LOAN",
        "PAN already linked to another open loan across ANY LSP → REJECTED with BORROWER_HAS_OPEN_LOAN (one-open-loan rule).",
        "LSP API Client", "Auto-Approval", "POST /api/v1/lsp/loan-applications",
        severity="Critical",
        notes="Global rule per project CONTEXT.md.")

    ec(27, "Duplicate intake (idempotency replay)",
        "POST /lsp/loan-applications with same Idempotency-Key + fingerprint → returns first response (200 with same applicationId); fingerprint mismatch → 409.",
        "LSP API Client", "Idempotency", "POST /api/v1/lsp/loan-applications",
        severity="High",
        notes="LspApiIdempotencyService claims via fingerprint.")

    ec(28, "Invalid payload — missing required field",
        "POST /lsp/loan-applications missing lspLoanId / productId / loanAmount → 400 with validation error array; no DB write.",
        "LSP API Client", "Validation", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    ec(29, "Invalid PAN format",
        "PAN not matching AAAAA9999A regex → 400 validation error.",
        "LSP API Client", "Validation", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    ec(30, "Mismatched lspId in payload vs token",
        "Request body's lspId ≠ authenticated LSP → AccessDeniedException (403).",
        "LSP API Client", "Tenancy / Isolation", "POST /api/v1/lsp/loan-applications",
        severity="Critical")

    ec(31, "Malformed JSON body",
        "Invalid JSON → 400 from Spring HttpMessageNotReadable handler; no stack trace leaked.",
        "Any client", "Validation", "Any POST/PUT/PATCH endpoint",
        severity="Low")

    ec(32, "Oversized request body",
        "JSON beyond Spring max-request-size → 413 / 400; documents > policy max → 400 with size violation.",
        "Any client", "Validation", "POST /lsp/loan-applications/{id}/documents",
        severity="Medium")

    # ---------- STATE MACHINE VIOLATIONS ----------
    ec(33, "Illegal status transition (REJECTED → anything)",
        "POST /ops/loan-applications/{id}/status-transitions from REJECTED → 400/409 because terminal status cannot transition (canTransitionTo returns false).",
        "System Administrator", "Lifecycle", "POST /api/v1/internal/ops/loan-applications/{id}/status-transitions",
        severity="High")

    ec(34, "Illegal status transition (CLOSED → anything)",
        "Same as EC-033 for CLOSED (terminal).",
        "System Administrator", "Lifecycle", "POST /status-transitions",
        severity="High")

    ec(35, "Illegal status transition (FORECLOSED → anything)",
        "FORECLOSED rejects all subsequent transitions.",
        "System Administrator", "Lifecycle", "POST /status-transitions",
        severity="High")

    ec(36, "Illegal status transition (INVALID → anything)",
        "INVALID rejects all transitions; cascaded loan_account stays INVALID-mirrored.",
        "System Administrator", "Lifecycle", "POST /status-transitions",
        severity="High")

    ec(37, "Invalidate after disbursement",
        "POST /lsp/loan-applications/{id}/invalid on DISBURSED/UNDER_REPAYMENT/CLOSED/FORECLOSED → 400 (only PRE_DISBURSAL_STATUSES allowed).",
        "LSP API Client / LSP_UI_WRITE", "Lifecycle", "POST /api/v1/lsp/loan-applications/{id}/invalid",
        severity="High")

    ec(38, "Invalidation idempotency — same key repeated",
        "Second call with identical Idempotency-Key + fingerprint returns first response (200, same applicationId/status); different fingerprint → 409.",
        "LSP API Client", "Idempotency", "POST /api/v1/lsp/loan-applications/{id}/invalid",
        severity="High")

    ec(39, "Manual status transition without required reason",
        "POST /status-transitions with reasonCode-required target but missing reasonCode → 400.",
        "System Administrator", "Lifecycle", "POST /status-transitions",
        severity="Medium")

    # ---------- DOCUMENTS ----------
    ec(40, "Upload disallowed MIME type",
        "Upload .exe / .bin → 400 from DocumentUploadPolicy; not persisted.",
        "LSP API Client / LSP_UI_WRITE", "Documents", "POST /api/v1/lsp/loan-applications/{id}/documents",
        severity="High")

    ec(41, "Upload file exceeding size cap",
        "PAN/AADHAAR upload > 5 MB → 400; other docs > 10 MB → 400 (DocumentUploadPolicy per-type caps).",
        "LSP API Client / LSP_UI_WRITE", "Documents", "POST .../documents",
        severity="Medium")

    ec(42, "Upload document not on checklist",
        "documentType not in LoanApplicationDocumentRequirements for that product → 400.",
        "LSP API Client", "Documents", "POST .../documents",
        severity="Medium")

    ec(43, "Upload to non-existent application",
        "POST .../documents on random UUID → 404.",
        "LSP API Client", "Documents", "POST .../documents",
        severity="Low")

    ec(44, "Cross-tenant document download",
        "OPS GET /kyc-documents/{type}/content for app outside scope → 404; download attempt audited.",
        "System Administrator / Operations User", "Documents", "GET /api/v1/internal/ops/loan-applications/{id}/kyc-documents/{type}/content",
        severity="Critical")

    ec(45, "Download non-existent document type",
        "GET .../kyc-documents/PAN_CARD/content on app with no PAN_CARD uploaded → 404 with checklist-pending reason.",
        "Operations User", "Documents", "GET .../kyc-documents/{type}/content",
        severity="Low")

    # ---------- DISBURSEMENT ----------
    ec(46, "Disbursement attempted on non-eligible status",
        "POST /disbursement-requests on AWAITING_APPROVAL or REJECTED or INVALID → 400 (only APPROVED_PENDING_DISBURSAL / DISBURSEMENT_RETRY allowed).",
        "System Administrator", "Disbursement", "POST /api/v1/internal/ops/loan-applications/{id}/disbursement-requests",
        severity="High")

    ec(47, "Disbursement attempted without repayment schedule",
        "Eligible status but no schedule on file → 400 with schedule-missing reason; LoanDisbursementWorker skips and emits OpsAlert.",
        "System Administrator / System", "Disbursement", "POST /disbursement-requests",
        severity="High")

    ec(48, "Disbursement attempted with bank mismatch",
        "Bank check failed earlier or borrower bank holder name fails matcher → disbursement blocked; LoanDisbursementBankMismatchLog row + ops alert.",
        "System / System Administrator", "Disbursement", "POST /disbursement-requests, worker",
        severity="High")

    ec(49, "Disbursement retry exhaustion",
        "After N retryable failures, application stays DISBURSEMENT_RETRY → OpsAlert DISBURSEMENT_RETRY_EXHAUSTED; manual SYSTEM_ADMIN intervention required.",
        "System (worker)", "Disbursement", "Worker: LoanDisbursementWorker",
        severity="High")

    ec(50, "Mock disbursement outside test mode",
        "POST /disbursement-requests/mock-outcome in env where mock adapter disabled → 404/400.",
        "System Administrator", "Disbursement (Mock)", "POST .../mock-outcome",
        severity="Low")

    # ---------- REPAYMENT ----------
    ec(51, "Payment without Idempotency-Key",
        "POST /payments missing Idempotency-Key header → 400 (Gap #17 contract enforces).",
        "System Administrator / Operations User / LSP API Client", "Servicing", "POST /api/v1/internal/ops/loan-applications/{id}/payments, POST /api/v1/lsp/loans/{id}/payments",
        severity="High")

    ec(52, "Payment without targetInstallmentId",
        "POST /payments missing targetInstallmentId → 400.",
        "System Administrator / Operations User / LSP API Client", "Servicing", "POST /payments",
        severity="High")

    ec(53, "Partial installment amount rejected",
        "Amount ≠ installment.due → 400 with BR-13 violation; no payment recorded.",
        "Any payer", "Servicing", "POST /payments",
        severity="Critical",
        notes="Project policy: partial installments not accepted.")

    ec(54, "Payment to already-PAID installment",
        "targetInstallmentId already PAID → 400/409 with installment-already-paid reason.",
        "Any payer", "Servicing", "POST /payments",
        severity="High")

    ec(55, "Payment idempotency replay",
        "Same Idempotency-Key + fingerprint → returns prior response; different fingerprint with same key → 409.",
        "Any payer", "Idempotency", "POST /payments",
        severity="High")

    ec(56, "Payment with invalid channel",
        "channel not in NEFT/RTGS/IMPS → 400 enum validation.",
        "Any payer", "Servicing", "POST /payments",
        severity="Medium")

    ec(57, "Payment on closed/foreclosed loan",
        "POST /payments on CLOSED or FORECLOSED loan → 400.",
        "Any payer", "Servicing", "POST /payments",
        severity="High")

    # ---------- FORECLOSURE ----------
    ec(58, "Foreclosure quote on non-eligible loan",
        "POST /foreclosure-quote on AWAITING_APPROVAL or REJECTED → 400.",
        "System Administrator / LSP API Client", "Servicing", "POST /api/v1/lsp/loans/{id}/foreclosure-quote",
        severity="High")

    ec(59, "Foreclosure execute with expired quote",
        "POST /foreclosure-quotes/{quoteId}/execute on quote past validity → 400/409.",
        "System Administrator", "Servicing", "POST .../foreclosure-quotes/{quoteId}/execute",
        severity="High")

    ec(60, "Foreclosure execute on closed/foreclosed loan",
        "Loan already CLOSED/FORECLOSED → 400.",
        "System Administrator", "Servicing", "POST .../foreclosure-quotes/{quoteId}/execute",
        severity="Medium")

    # ---------- WEBHOOKS ----------
    ec(61, "Webhook to invalid URL (DNS failure)",
        "Subscriber URL unresolvable → RETRYABLE_FAILURE with attempt counter increment; backoff per WebhookOutboxProperties.",
        "System (worker)", "Webhooks", "Worker: WebhookOutboxDispatchWorker",
        severity="Medium")

    ec(62, "Webhook subscriber returns 5xx",
        "HTTP 5xx → RETRYABLE_FAILURE; honor backoff.",
        "System (worker)", "Webhooks", "Worker: WebhookOutboxDispatchWorker",
        severity="Medium")

    ec(63, "Webhook subscriber returns 4xx (non-retryable)",
        "HTTP 4xx (except 408/429) → PERMANENT_FAILURE; OpsAlert WEBHOOK_DEAD_LETTER raised; available for manual redrive.",
        "System (worker)", "Webhooks", "Worker: WebhookOutboxDispatchWorker",
        severity="High")

    ec(64, "Webhook redrive without permission",
        "Non-SYSTEM_ADMIN POST /webhook-outbox/{id}/redrive → 403.",
        "Operations User", "Webhooks", "POST /api/v1/internal/admin/webhook-outbox/{id}/redrive",
        severity="High")

    ec(65, "Webhook signature header missing on payload",
        "Subscriber test should verify X-Bhawana-Signature is present and HMAC matches; absence indicates regression.",
        "External LSP webhook consumer", "Webhooks", "Webhook payload",
        severity="High")

    ec(66, "SSRF in webhook URL",
        "Subscription URL pointing to internal hosts (169.254.169.254, localhost) → rejected by SsrfSafeUrlValidator at save time.",
        "System Administrator", "Security/Webhooks", "PUT /api/v1/internal/admin/lsps/{lspId}/webhook-subscription",
        severity="Critical")

    # ---------- REPORTS ----------
    ec(67, "MIS preview leaks unmasked aadhaar",
        "GET /reports/portfolio-mis/preview MUST mask aadhaar as XXXXXXXX#### and bank_account as XXXX#### per Gap #10; any unmasked → Critical fail.",
        "System Administrator", "Reports / PII", "GET /api/v1/internal/reports/portfolio-mis/preview",
        severity="Critical")

    ec(68, "MIS CSV download large dataset",
        "Async path: POST /reports/portfolio-mis/requests for year-wide range → request transitions PENDING→PROCESSING→COMPLETED; CSV downloadable.",
        "System Administrator", "Reports", "POST /reports/portfolio-mis/requests, GET /requests/{id}/download",
        severity="Medium")

    ec(69, "Report download unauthorized user",
        "OPS_USER hitting /reports/* → 403.",
        "Operations User", "Reports", "GET /api/v1/internal/reports/*",
        severity="High")

    ec(70, "Empty filter range — no data",
        "MIS request with future-only date range → empty CSV with header row only; no 500.",
        "System Administrator", "Reports", "GET /reports/portfolio-mis",
        severity="Low")

    # ---------- BORROWER & TENANT ISOLATION ----------
    ec(71, "Borrower lookup across tenants (LSP)",
        "LSP A querying borrower belonging only to LSP B → 404 (not 403).",
        "LSP API Client", "Tenancy", "GET /api/v1/lsp/borrowers/{id}/bank-details",
        severity="Critical")

    ec(72, "Borrower bank update with invalid IFSC",
        "PATCH bank-details with malformed IFSC → 400.",
        "LSP API Client / SYSTEM_ADMIN", "Borrowers / Validation", "PATCH /api/v1/lsp/borrowers/{id}/bank-details",
        severity="Medium")

    ec(73, "Borrower aadhaar masked on read",
        "GET /admin/borrowers/{id} returns aadhaar masked (XXXXXXXX####); raw aadhaar never returned in any API.",
        "Internal roles", "PII Masking", "GET /api/v1/internal/admin/borrowers/{id}",
        severity="Critical",
        notes="Gap #1 — PII reveal endpoint removed; mask everywhere.")

    # ---------- RATE LIMITING ----------
    ec(74, "Rate limit on /auth/login per IP+username",
        "Excess requests in window → 429 with Retry-After; RateLimitFilter emits ops alert on breach (Follow-up #2).",
        "Any client", "Rate Limiting", "POST /api/v1/auth/login",
        severity="High")

    ec(75, "Rate limit on /auth/token per client_id",
        "Excess token issuance → 429.",
        "LSP API Client", "Rate Limiting", "POST /api/v1/auth/token",
        severity="Medium")

    ec(76, "Rate limit on /lsp/loan-applications per LSP",
        "Burst of intake calls beyond bucket → 429 with Retry-After.",
        "LSP API Client", "Rate Limiting", "POST /api/v1/lsp/loan-applications",
        severity="Medium")

    # ---------- ALERTS ----------
    ec(77, "Acknowledge non-existent alert",
        "POST /alerts/{random-uuid}/acknowledge → 404.",
        "Operations User", "Alerts", "POST /api/v1/internal/alerts/{id}/acknowledge",
        severity="Low")

    ec(78, "Acknowledge note over 500 chars",
        "Body.note length > 500 → 400 (Gap #15 validation).",
        "Operations User", "Alerts / Validation", "POST /alerts/{id}/acknowledge",
        severity="Low")

    ec(79, "Acknowledge already-acknowledged alert",
        "Second ack on same alert → idempotent (returns updated ack metadata; no error).",
        "Operations User", "Alerts", "POST /alerts/{id}/acknowledge",
        severity="Low")

    ec(80, "Alert dedupe (same condition, same loan)",
        "AlertRuleSchedulerWorker re-evaluates and finds same condition → createAlertIfAbsent returns existing open alert (no duplicate row).",
        "System (worker)", "Alerts", "Worker: AlertRuleSchedulerWorker",
        severity="Medium")

    # ---------- FRONTEND / UX ----------
    ec(81, "LSP_UI_WRITE attempts /loan-applications (internal route)",
        "RequireInternal guard blocks → redirected to /my-loans.",
        "LSP UI User", "Frontend Guards", "/loan-applications UI",
        severity="High",
        test_type="UI")

    ec(82, "OPS_USER navigates to /reports",
        "RequireRole guard blocks (SYSTEM_ADMIN_ONLY); shows PermissionDeniedState.",
        "Operations User", "Frontend Guards", "/reports UI",
        severity="High",
        test_type="UI")

    ec(83, "Browser refresh on detail page preserves session",
        "F5 on /loan-applications/{id} → silent refresh + same view; no auth bounce.",
        "Internal user", "Frontend / Session", "/loan-applications/{id}",
        severity="Medium",
        test_type="UI")

    ec(84, "401 from backend mid-session",
        "FE intercepts 401 → triggers /auth/refresh → retries; if both fail → redirect to /login with return URL.",
        "Internal user", "Frontend / Session", "Any UI page",
        severity="High",
        test_type="UI",
        notes="Memory: #89 live backend 401-loop regression — TestTenantContextRestoreFilter is @Profile('test') only; admin tenant scope not set on prod chain. Likely Fails.")

    ec(85, "Console errors on screen load",
        "Each major screen loads without uncaught errors / unhandled promise rejections in browser console.",
        "Internal user", "Frontend", "All UI pages",
        severity="Medium",
        test_type="UI")

    ec(86, "Failed network request surfaces ErrorState",
        "API 5xx → ErrorState component with retry CTA; not a blank screen.",
        "Internal user", "Frontend", "All data-fetching pages",
        severity="Medium",
        test_type="UI")

    ec(87, "Empty list states",
        "/loan-applications, /borrowers, /alerts, /reports with no rows → EmptyState component (not bare table).",
        "Internal user", "Frontend / UX", "All list pages",
        severity="Low",
        test_type="UI")

    ec(88, "Pagination boundaries",
        "Hit first/last page; ensure prev/next disabled correctly; page count consistent with total when paginationDetails=true.",
        "Internal user", "Frontend / List", "/loan-applications, /borrowers, /audit, etc.",
        severity="Medium",
        test_type="UI + API")

    ec(89, "Filter combinations on /loan-applications",
        "Status + LSP + date range simultaneously → server query honors all; result count reconciles with cleared-filter baseline.",
        "Operations User", "Frontend / Filtering", "/loan-applications",
        severity="Medium",
        test_type="UI + API")

    ec(90, "Sensitive-data toggle",
        "Toggling reveals masked columns; click is audited; refresh respects last toggle state (or resets, depending on spec).",
        "Operations User", "Frontend / PII", "/loan-applications",
        severity="High",
        test_type="UI")

    # ---------- AUDIT ----------
    ec(91, "Audit search with since > until",
        "GET /audit-events?since=X&until=Y where X>Y → 400.",
        "System Administrator", "Audit / Validation", "GET /api/v1/internal/admin/audit-events",
        severity="Low")

    ec(92, "Audit limit clamping",
        "limit=1000 → clamped to 500; limit=0/-1 → reset to default 100.",
        "System Administrator", "Audit", "GET /api/v1/internal/admin/audit-events",
        severity="Low")

    ec(93, "Audit unknown stream value",
        "streams=BOGUS → 400.",
        "System Administrator", "Audit / Validation", "GET /api/v1/internal/admin/audit-events",
        severity="Low")

    ec(94, "Audit INTAKE detail masks aadhaar in payload",
        "Even when intake JSON contained raw aadhaar, response detail masks XXXXXXXX####.",
        "System Administrator", "Audit / PII", "GET /api/v1/internal/admin/audit-events",
        severity="Critical")

    ec(95, "Document access logged on every download",
        "Each GET .../kyc-documents/{type}/content writes a LoanApplicationDocumentAccessAudit row visible in Audit Explorer.",
        "Operations User", "Audit / Compliance", "GET .../kyc-documents/{type}/content",
        severity="High")

    # ---------- CONSISTENCY / DATA ----------
    ec(96, "Status displayed in UI matches DB",
        "After manual override or auto-rejection, /loan-applications list + detail page reflect new status without manual refresh (react-query invalidate).",
        "Internal user", "Consistency", "/loan-applications/{id}",
        severity="Medium",
        test_type="UI + DB")

    ec(97, "Webhook event count in detail matches outbox",
        "/loan-applications/{id} Webhooks tab count matches GET .../webhook-events length (cap 200).",
        "Operations User", "Consistency", "/loan-applications/{id}",
        severity="Medium")

    ec(98, "First-payment transition disbursed→under_repayment",
        "Recording first installment payment on DISBURSED loan flips status to UNDER_REPAYMENT immediately; subsequent payments stay UNDER_REPAYMENT.",
        "Operations User", "Servicing / State Machine", "POST /payments",
        severity="High")

    ec(99, "Last-installment transition under_repayment→closed",
        "Final installment paid → loan closes (CLOSED) and fully-repaid webhook queued.",
        "Operations User", "Servicing / State Machine", "POST /payments",
        severity="High")

    ec(100, "Rejection reason persisted in transition row",
         "Auto-rejection at AWAITING_APPROVAL writes failed RuleCode array into loan_application_status_transition.rejection_reason_json (Gap #11).",
         "System (engine)", "Auto-Approval / Audit", "DB: loan_application_status_transition.rejection_reason_json",
         severity="Medium")

    ec(101, "PAN dedup links to existing borrower",
         "Second LSP submits app for same PAN → borrower row reused (no duplicate); cross-tenant access controlled.",
         "LSP API Client", "Borrowers / Dedup", "POST /api/v1/lsp/loan-applications",
         severity="High")

    ec(102, "Local backend 401 loop regression (#89 memory)",
         "Authenticated request to admin endpoint with valid JWT should succeed in prod-like profile; if filter chain runs before MVC interceptor sets admin scope (per memory), every call 401s.",
         "System Administrator", "Tenancy / Filters", "Any /api/v1/internal/admin/* on live backend",
         severity="Critical",
         notes="Per memory project_89_prod_regression: TestTenantContextRestoreFilter is @Profile('test') only. AdminTenantDataAccessFilter is newly added (untracked) — verify whether it now closes the gap.")

    ec(103, "Processing fee deduction at disbursement (ADR 0004)",
         "Disbursed amount = principal - processing_fee per product config; loan_account.disbursed_amount + audit reflect deduction.",
         "System", "Disbursement / Product", "POST /disbursement-requests, LoanDisbursementService",
         severity="High",
         notes="Per docs/adr/0004-processing-fee-deduction.md.")

    ec(104, "Loan account creation only on approval",
         "loan_account row created exclusively when application transitions to APPROVED_PENDING_DISBURSAL; never before; INVALID/REJECTED apps have no loan_account.",
         "System", "Origination", "DB: loan_account",
         severity="High")

    ec(105, "Webhook ordering per aggregate",
         "Events for one loan are delivered in created_at order; out-of-order subscriber receives → spec/backend should preserve sequence.",
         "External LSP webhook consumer", "Webhooks", "WebhookOutboxDispatchWorker",
         severity="Medium")

    ec(106, "Bank holder-name fuzzy match",
         "BankAccountHolderNameMatcher handles minor case/space/initials variations; gross mismatch → mismatch logged.",
         "System", "Disbursement / Validation", "BankAccountHolderNameMatcher",
         severity="Medium")

    ec(107, "Schedule submission validation",
         "PUT /repayment-schedule with sum of installments ≠ principal+interest → 400 with ScheduleViolationType detail.",
         "LSP API Client", "Disbursement / Validation", "PUT /api/v1/lsp/loan-applications/{id}/repayment-schedule",
         severity="High")

    ec(108, "Schedule update after disbursement blocked",
         "Re-submitting schedule for DISBURSED loan → 400 (schedule locked post-disbursal).",
         "LSP API Client", "Servicing", "PUT /repayment-schedule",
         severity="High")

    ec(109, "Concurrent disbursement attempts",
         "Two simultaneous disbursement requests for same loan → only one succeeds (DB row lock / status precondition); second sees 409 / no-op.",
         "System Administrator / System (worker)", "Disbursement / Concurrency", "POST /disbursement-requests + worker race",
         severity="High")

    ec(110, "Audit explorer correlationId trace",
         "Given a correlationId, FE post-filter (and future BE filter) returns full cross-domain timeline of that operation.",
         "System Administrator", "Audit", "GET /api/v1/internal/admin/audit-events",
         severity="Medium")

    return rows


def main() -> None:
    wb = Workbook()
    ws_uc = wb.active
    ws_uc.title = "Use Cases"
    ws_ec = wb.create_sheet("Edge Cases")

    uc_rows = use_case_rows()
    ec_rows = edge_case_rows()

    write_rows(ws_uc, uc_rows)
    style_sheet(ws_uc, len(uc_rows))

    write_rows(ws_ec, ec_rows)
    style_sheet(ws_ec, len(ec_rows))

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}: {len(uc_rows)} use cases, {len(ec_rows)} edge cases.")


if __name__ == "__main__":
    main()
