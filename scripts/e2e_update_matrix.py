#!/usr/bin/env python3
"""Update e2e-test-matrix.xlsx row(s) after test execution."""
import argparse
import sys
from pathlib import Path

import openpyxl

MATRIX = Path(__file__).resolve().parent.parent / "e2e-test-matrix.xlsx"
COL_STATUS = 4
COL_ACTUAL = 5
COL_STEPS = 11
COL_NOTES = 13


def find_row(ws, test_id: str) -> int | None:
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == test_id:
            return r
    return None


def update(test_id: str, status: str, actual: str, steps: str = "", notes: str = ""):
    wb = openpyxl.load_workbook(MATRIX)
    for sheet_name in ("Use Cases", "Edge Cases"):
        ws = wb[sheet_name]
        row = find_row(ws, test_id)
        if row:
            ws.cell(row, COL_STATUS, status)
            ws.cell(row, COL_ACTUAL, actual)
            if steps:
                ws.cell(row, COL_STEPS, steps)
            if notes:
                existing = ws.cell(row, COL_NOTES).value or ""
                ws.cell(row, COL_NOTES, f"{existing}; {notes}".strip("; "))
            wb.save(MATRIX)
            print(f"Updated {test_id} on '{sheet_name}' row {row} -> {status}")
            return
    print(f"WARNING: {test_id} not found", file=sys.stderr)
    sys.exit(1)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("test_id")
    p.add_argument("status", choices=["Pass", "Fail", "Blocked", "Not Applicable"])
    p.add_argument("actual")
    p.add_argument("--steps", default="")
    p.add_argument("--notes", default="")
    args = p.parse_args()
    update(args.test_id, args.status, args.actual, args.steps, args.notes)


if __name__ == "__main__":
    main()
