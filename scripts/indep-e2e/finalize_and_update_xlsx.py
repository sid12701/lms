"""Patch ambiguous results to their resolved verdicts, then write ALL gap-test
rows into e2e-test-matrix.xlsx (update existing UC/EC rows; append new rows).
"""
from __future__ import annotations
import json
from pathlib import Path
import openpyxl

R = {r["id"]: r for r in json.load(open("gap_results.json"))}


def patch(id_, **kw):
    if id_ in R:
        R[id_].update(kw)


# EC-063: corrected — hard 4xx (400) => PERMANENT_FAILURE (404/410 are intentionally soft-retryable)
patch("EC-063", status="Pass",
      actual="Hard 4xx (httpbin 400) -> permanentFailures=1 (DEAD_LETTERED). Note: 404/410 are intentionally classified SOFT (retryable) by WebhookOutboxDispatchExecutor.classify().",
      notes="Independent re-run with public receiver. Webhook failure classification: 2xx=DELIVERED, 5xx/408/429=RETRYABLE, 404/410=soft RETRYABLE, other 4xx=PERMANENT. Correct/by-design.")

# EC-048: confirmed BUG B3
patch("EC-048", status="Fail", severity="High",
      actual="disbursement-bank-check returns 500 INTERNAL_SERVER_ERROR on ANY mismatch (name or account). MATCH path returns 200 {status:OK}. Should be 422 DISBURSEMENT_VALIDATION_FAILED with the mismatch surfaced.",
      notes="NEW BUG B3. Root: failure branch calls recordHardDisbursementBankMismatch() (writes bank_mismatch_log + ops alert) inside the LSP tenant transaction; the write throws -> 500. Match path (no write) returns 200. Partners get opaque 500 exactly when details mismatch.")

patch("EC-106", status="Fail", severity="Medium",
      actual="Same 500 as EC-048 (mismatch path crashes); fuzzy-match tolerance could not be evaluated because any non-exact name triggers the 500.",
      notes="Blocked by BUG B3 (bank-check 500 on mismatch). Re-test fuzzy tolerance after B3 fix.")

# EC-107: confirmed 500 on invalid provided schedule (same family as B3)
patch("EC-107", status="Fail", severity="High",
      actual="PUT /repayment-schedule mode=LSP_PROVIDED with sum!=principal on an APPROVED loan -> 500 INTERNAL_SERVER_ERROR (expected 4xx REPAYMENT_SCHEDULE_INVALID with ScheduleViolationType). GENERATED mode works fine.",
      notes="NEW BUG (B3 family). Validation detects the violation but the catch-branch emits an ops alert (emitLspProvidedScheduleViolation) inside the LSP tenant tx -> write throws -> 500 instead of the structured 422.")

# EC-003: honest framing — brute-force blocked but lockout vs rate-limit not isolable at /auth/login
patch("EC-003", status="Pass", severity="High",
      actual="Repeated wrong logins are blocked: codes interleave 401 then 429; correct password afterwards still blocked (429). Brute-force is prevented.",
      notes="Account-lockout (V94) vs IP rate-limit (10/min) cannot be cleanly isolated at /auth/login since both gate the same path. Security goal (brute-force prevented) is met. To unit-verify lockout specifically, disable rate-limit in a test profile.")

# ---- write to workbook ----
ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "e2e-test-matrix.xlsx"
wb = openpyxl.load_workbook(XLSX)


def sheet_for(id_):
    return wb["Use Cases"] if id_.startswith("UC") else wb["Edge Cases"]


def find_row(ws, id_):
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, 1).value or "") == id_:
            return r
    return None


COLS = ["id", "title", "expected", "status", "actual", "actor", "module", "source", "ttype", "api", "steps", "severity", "notes"]
updated, appended = [], []

for id_, row in sorted(R.items()):
    ws = sheet_for(id_)
    rn = find_row(ws, id_)
    vals = [row.get(k, "") for k in COLS]
    if rn:
        # update result columns (status, actual, steps, severity, notes); keep descriptive cols unless blank
        ws.cell(rn, 4, row.get("status", ""))
        ws.cell(rn, 5, row.get("actual", ""))
        if row.get("steps"):
            ws.cell(rn, 11, row.get("steps"))
        if row.get("severity"):
            ws.cell(rn, 12, row.get("severity"))
        if row.get("notes"):
            ws.cell(rn, 13, row.get("notes"))
        # fill descriptive cols if currently empty
        for ci, key in [(6, "actor"), (7, "module"), (8, "source"), (9, "ttype"), (10, "api")]:
            if not (ws.cell(rn, ci).value or "").strip() and row.get(key):
                ws.cell(rn, ci, row.get(key))
        updated.append(id_)
    else:
        rn = ws.max_row + 1
        for ci, key in enumerate(COLS, start=1):
            ws.cell(rn, ci, row.get(key, ""))
        appended.append(id_)

wb.save(XLSX)
print(f"Updated {len(updated)} existing rows: {updated}")
print(f"Appended {len(appended)} new rows: {appended}")
