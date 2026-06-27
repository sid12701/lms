"""Definitive: remap misaligned fields, apply resolved verdicts, rewrite matrix cleanly."""
from __future__ import annotations
import json
from pathlib import Path
import openpyxl

SEV = {"Critical", "High", "Medium", "Low"}
raw = json.load(open("gap_results.json"))

# ---- 1. remap fields: api->steps(action), steps->severity(word), severity->notes(text) ----
fixed = {}
for r in raw:
    a, b, c = r.get("steps", ""), r.get("severity", ""), r.get("notes", "")
    sevword = next((x for x in (a, b, c) if x in SEV), "")
    note = next((x for x in (c, b, a) if x and x not in SEV), "")
    fixed[r["id"]] = {
        "id": r["id"],
        "title": r["title"],
        "expected": r["expected"],
        "status": r["status"],
        "actual": r["actual"],
        "actor": r["actor"],
        "module": r["module"],
        "source": r.get("source", "Independent E2E"),
        "ttype": r.get("ttype", "API"),
        "api": r.get("api", ""),          # action/endpoint (col10 for new rows; existing keep sheet)
        "steps": r.get("api", ""),         # steps performed (col11)
        "severity": sevword,               # col12
        "notes": note,                     # col13
    }

# ---- 2. apply resolved verdicts/wording (post-diagnosis) ----
PATCH = {
    "EC-063": dict(status="Pass", severity="High",
        actual="Hard 4xx (httpbin 400) -> permanentFailures=1 (DEAD_LETTERED). 404/410 are intentionally classified SOFT (retryable).",
        notes="Independent re-run with public receiver (httpbin). Classifier: 2xx=DELIVERED, 5xx/408/429=RETRYABLE, 404/410=soft RETRYABLE, other 4xx=PERMANENT (by design)."),
    "EC-048": dict(status="Fail", severity="High",
        actual="disbursement-bank-check -> 500 INTERNAL_SERVER_ERROR on ANY mismatch (name or account); MATCH path returns 200 {status:OK}. Expected 422 DISBURSEMENT_VALIDATION_FAILED.",
        notes="NEW BUG B3. Failure branch calls recordHardDisbursementBankMismatch() (writes bank_mismatch_log + ops alert) inside the LSP tenant tx -> write throws -> 500. Partners get opaque 500 exactly on mismatch."),
    "EC-106": dict(status="Fail", severity="Medium",
        actual="Same 500 as EC-048 on any non-exact name; fuzzy-match tolerance not evaluable until B3 fixed.",
        notes="Blocked by BUG B3 (bank-check 500 on mismatch). Re-test fuzzy tolerance after fix."),
    "EC-107": dict(status="Fail", severity="High",
        actual="PUT /repayment-schedule mode=LSP_PROVIDED with bad sum on APPROVED loan -> 500 (expected 4xx REPAYMENT_SCHEDULE_INVALID). GENERATED mode works.",
        notes="NEW BUG (B3 family). Validation detects the violation, then emitLspProvidedScheduleViolation() writes an ops alert inside the LSP tenant tx -> 500 instead of structured 422."),
    "EC-003": dict(status="Pass", severity="High",
        actual="Repeated wrong logins blocked: codes interleave 401 then 429; correct password afterwards still blocked. Brute-force prevented.",
        notes="Account-lockout (V94) vs IP rate-limit (10/min) not cleanly isolable at /auth/login (shared path). Security goal met. Unit-verify lockout with rate-limit disabled in a test profile."),
    "EC-103": dict(status="Fail", severity="High",
        actual="NOT IMPLEMENTED. MIS disbursalAmount=150000 (full principal); processingFeeAmount=2250 is synthetic. No deduction, no persisted fee column.",
        notes="ADR 0004 (Proposed) calls this 'a fiction'. No LoanFeeCalculator, no processing_fee_amount column; LoanAccountRepository disbursedAmount=principalAmount. Borrower receives full principal."),
    "EC-112": dict(status="Fail", severity="High",
        actual="Borrower-360 field 'bankAccountNumberMasked' returns RAW account (e.g. 256335623472); MIS CSV 'Bank Account Number' + INTAKE audit also raw; UI Banking section shows it raw. Aadhaar IS masked.",
        notes="NEW BUG B1. BorrowerAdminController:139 passes raw getBankAccountNumber() into the *Masked field; no maskBankAccount() helper. Missed by EC-067 (only scanned aadhaar)."),
    "EC-113": dict(status="Fail", severity="High",
        actual="Aadhaar masked everywhere (pass). Bank account RAW in borrower-360, MIS preview/CSV, INTAKE audit, and UI. Full-body regex finds the unmasked account.",
        notes="NEW standing regression test: regex full JSON/CSV for 12-digit (aadhaar) AND 9-18-digit (account). Catches B1 that the per-key-name check missed."),
    "EC-114": dict(status="Blocked", severity="High",
        actual="Cannot create an overdue installment: schedule due dates are future and no business-date override exists. New disbursed loans show delinquencyBucket=CURRENT.",
        notes="NEW: needs a business-clock override or seeded overdue loan to exercise DPD bucketing math (dashboard + borrower aggregate + MIS)."),
}
for id_, kw in PATCH.items():
    if id_ in fixed:
        fixed[id_].update(kw)

# ---- 3. write to workbook ----
ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "e2e-test-matrix.xlsx"
wb = openpyxl.load_workbook(XLSX)
COLS = ["id", "title", "expected", "status", "actual", "actor", "module", "source", "ttype", "api", "steps", "severity", "notes"]


def ws_for(id_):
    return wb["Use Cases"] if id_.startswith("UC") else wb["Edge Cases"]


def find(ws, id_):
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, 1).value or "") == id_:
            return r
    return None


upd, app = [], []
for id_, row in sorted(fixed.items()):
    ws = ws_for(id_)
    rn = find(ws, id_)
    if rn:
        ws.cell(rn, 4, row["status"])
        ws.cell(rn, 5, row["actual"])
        ws.cell(rn, 11, row["steps"])
        ws.cell(rn, 12, row["severity"])
        ws.cell(rn, 13, row["notes"])
        for ci, k in [(6, "actor"), (7, "module"), (8, "source"), (9, "ttype"), (10, "api")]:
            if not (ws.cell(rn, ci).value or "").strip():
                ws.cell(rn, ci, row[k])
        upd.append(id_)
    else:
        rn = ws.max_row + 1
        for ci, k in enumerate(COLS, 1):
            ws.cell(rn, ci, row[k])
        app.append(id_)

# EC-067 correction (false-pass)
ec = wb["Edge Cases"]
for r in range(2, ec.max_row + 1):
    if ec.cell(r, 1).value == "EC-067":
        ec.cell(r, 4, "Fail")
        ec.cell(r, 5, "Aadhaar IS masked, but BANK ACCOUNT NUMBER is RAW in MIS preview/CSV (+ borrower-360 + INTAKE audit). EC-067 requires bank_account masked as XXXX####.")
        ec.cell(r, 12, "High")
        ec.cell(r, 13, "FALSE PASS in prior run: phase9_data_adr.py only scanned aadhaar-keyed fields, never bank account. See EC-112/EC-113, bug B1.")

wb.save(XLSX)
print(f"Updated {len(upd)}: {upd}")
print(f"Appended {len(app)}: {app}")
